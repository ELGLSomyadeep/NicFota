from logging import root
import sys
import os
import re
import json
import time
import socket
import platform
import subprocess
import threading
from datetime import datetime
from urllib.parse import urlparse
import urllib.request
import urllib.error

from PySide6 import QtCore, QtGui, QtWidgets
from PySide6.QtCore import Qt

# ─────────────────────────────────────────────────────────────
#  Settings persistence
# ─────────────────────────────────────────────────────────────
if getattr(sys, "frozen", False):
    SETTINGS_FILE = os.path.join(os.path.dirname(sys.executable), "fota_settings.json")
else:
    SETTINGS_FILE = os.path.join(os.path.dirname(__file__), "fota_settings.json")


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def load_settings() -> dict:
    try:
        with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_settings(data: dict):
    try:
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
    except Exception:
        pass


# ─────────────────────────────────────────────────────────────
#  Theme / palette  —  a bit brighter / more "alive" than the CTk version
# ─────────────────────────────────────────────────────────────
BRAND        = "#6fae2e"
BRAND_HOVER  = "#5c9424"
BRAND_LIGHT  = "#eef6e4"

DANGER       = "#e0433a"
DANGER_HOVER = "#c73a30"

WARNING      = "#f39c12"
WARNING_HOVER = "#d68910"

INFO         = "#2f8fd6"
INFO_HOVER   = "#2477b3"

SUCCESS      = "#27ae60"
SUCCESS_HOVER = "#219150"

SLATE        = "#8592a3"
SLATE_HOVER  = "#6f7c8d"

PURPLE       = "#8e6fd1"
PURPLE_HOVER = "#7a5cc0"

BG_MAIN      = "#f4f7f2"
CARD_BG      = "#ffffff"
CARD_BORDER  = "#e4e8de"
MUTED_TEXT   = "#8a8f86"
TEXT_DARK    = "#2b2f27"

FONT_FAMILY  = "Segoe UI"
MONO_FAMILY  = "Consolas"


def qfont(family=FONT_FAMILY, size=11, weight=QtGui.QFont.Normal, italic=False):
    f = QtGui.QFont(family, size)
    f.setWeight(weight)
    f.setItalic(italic)
    return f


FONT_H1    = lambda: qfont(FONT_FAMILY, 19, QtGui.QFont.Bold)
FONT_H2    = lambda: qfont(FONT_FAMILY, 13, QtGui.QFont.Bold)
FONT_LABEL = lambda: qfont(FONT_FAMILY, 11, QtGui.QFont.DemiBold)
FONT_BODY  = lambda: qfont(FONT_FAMILY, 11)
FONT_SMALL = lambda: qfont(FONT_FAMILY, 10)
FONT_MONO  = lambda: qfont(MONO_FAMILY, 10)


# ─────────────────────────────────────────────────────────────
#  Thread → UI-thread dispatcher (mirrors Tk's `self.after(0, fn, *a)`)
# ─────────────────────────────────────────────────────────────
class _Dispatcher(QtCore.QObject):
    fire = QtCore.Signal(object, tuple, dict)

    def __init__(self):
        super().__init__()
        self.fire.connect(self._run, QtCore.Qt.QueuedConnection)

    def _run(self, fn, args, kwargs):
        fn(*args, **kwargs)


# ─────────────────────────────────────────────────────────────
#  tkinter.messagebox-compatible shim, backed by QMessageBox
# ─────────────────────────────────────────────────────────────
class mb:
    @staticmethod
    def showerror(title, message):
        box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Critical, title, message)
        box.setStyleSheet(_MSGBOX_QSS)
        box.exec()

    @staticmethod
    def showinfo(title, message):
        box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Information, title, message)
        box.setStyleSheet(_MSGBOX_QSS)
        box.exec()

    @staticmethod
    def askyesno(title, message):
        box = QtWidgets.QMessageBox(QtWidgets.QMessageBox.Question, title, message,
                                     QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        box.setStyleSheet(_MSGBOX_QSS)
        return box.exec() == QtWidgets.QMessageBox.Yes


_MSGBOX_QSS = f"""
QMessageBox {{ background-color: {CARD_BG}; }}
QMessageBox QLabel {{ color: {TEXT_DARK}; font-size: 11pt; }}
QPushButton {{
    background-color: {BRAND}; color: white; border: none;
    border-radius: 7px; padding: 7px 18px; font-weight: 600; min-width: 70px;
}}
QPushButton:hover {{ background-color: {BRAND_HOVER}; }}
"""


# ─────────────────────────────────────────────────────────────
#  Tiny StringVar shim (tkinter-style .get()/.set(), Qt-signal driven)
# ─────────────────────────────────────────────────────────────
class StringVar(QtCore.QObject):
    changed = QtCore.Signal(str)

    def __init__(self, value=""):
        super().__init__()
        self._value = value

    def get(self):
        return self._value

    def set(self, value):
        self._value = value
        self.changed.emit(value)


# ─────────────────────────────────────────────────────────────
#  IPv6 socket helpers  (unchanged from the CustomTkinter version)
# ─────────────────────────────────────────────────────────────
def resolve_ipv6_sockaddr(ip: str, port: int):
    infos = socket.getaddrinfo(ip, port, socket.AF_INET6, socket.SOCK_STREAM)
    return infos[0][4]


def make_tcp_socket(timeout=5):
    s = socket.socket(socket.AF_INET6, socket.SOCK_STREAM)
    s.setsockopt(socket.IPPROTO_IPV6, socket.IPV6_V6ONLY, 0)
    s.settimeout(timeout)
    return s


def make_udp_socket(timeout=3):
    s = socket.socket(socket.AF_INET6, socket.SOCK_DGRAM)
    s.setsockopt(socket.IPPROTO_IPV6, socket.IPV6_V6ONLY, 0)
    s.settimeout(timeout)
    return s


def validate_ipv6_format(ip: str):
    ip = (ip or "").strip()
    if not ip:
        return False, "NIC IPv6 address field is empty."
    try:
        socket.inet_pton(socket.AF_INET6, ip)
        return True, "Valid IPv6 literal address."
    except OSError:
        pass
    try:
        socket.getaddrinfo(ip, None, socket.AF_INET6)
        return True, "Hostname resolves to a valid IPv6 address."
    except socket.gaierror:
        return False, "\"" + ip + "\" is not a valid IPv6 address or resolvable hostname."
    except Exception as e:
        return False, "IPv6 validation error: " + type(e).__name__ + ": " + str(e)


# ─────────────────────────────────────────────────────────────
#  NIC Firmware reader — HDLC frame sequence (unchanged logic)
# ─────────────────────────────────────────────────────────────
def printable_ascii(data: bytes) -> str:
    return ''.join(chr(b) if 32 <= b <= 126 else '.' for b in data)


def decode_firmware_response(rx: bytes) -> dict:
    result = {
        "raw_hex":      rx.hex(' ').upper(),
        "ascii":        printable_ascii(rx),
        "imei":         None,
        "firmware":     None,
        "manufacturer": None,
    }

    matches = re.findall(rb"\d{15}", rx)
    if matches:
        result["imei"] = matches[0].decode()

    idx = rx.find(b"HW")
    if idx != -1:
        fw = bytearray()
        while idx < len(rx):
            c = rx[idx]
            if 32 <= c <= 126:
                fw.append(c)
            else:
                break
            idx += 1
        result["firmware"] = fw.decode()

    if b"Landis+Gyr" in rx:
        start = rx.find(b"Landis+Gyr")
        mfr = bytearray()
        while start < len(rx):
            c = rx[start]
            if 32 <= c <= 126:
                mfr.append(c)
            else:
                break
            start += 1
        result["manufacturer"] = mfr.decode()

    return result


def _safe_close_socket(sock):
    if sock is None:
        return
    try:
        sock.shutdown(socket.SHUT_RDWR)
    except OSError:
        pass
    try:
        sock.close()
    except Exception:
        pass


HDLC_FRAMES = [
    ("SNRM",
     "7E A0 23 00 02 04 01 FD 93 56 61 81 80 14 05 02 27 0F 06 02 27 0F 07 04 00 00 00 01 08 04 00 00 00 01 BA E4 7E"),

    ("AARQ",
     "7E A0 4F 00 02 04 01 FD 10 DA C2 E6 E6 00 60 3E A1 09 06 07 60 85 74 05 08 01 01 8A 02 07 80 8B 07 60 85 74 05 08 02 02 AC 12 80 10 31 31 31 31 31 31 31 31 31 31 31 31 31 31 31 31 BE 10 04 0E 01 00 00 00 06 5F 1F 04 00 1C FF 3F 27 0F AD 12 7E"),

    ("RLRQ",
     "7E A0 2E 00 02 04 01 FD 32 9A FB E6 E6 00 C3 01 C1 00 0F 00 00 28 00 00 FF 01 01 09 10 FC 91 AE 8C 11 D3 8D 3E 50 19 91 CF D4 30 79 AF D3 D1 7E"),

    ("NIC Firmware",
     "7E A0 1C 00 02 04 01 FD 54 5B 1C E6 E6 00 C0 01 C1 00 01 00 87 64 00 00 FF 02 00 64 FA 7E"),

    ("DISC",
     "7E A0 0A 00 02 04 01 FD 53 E0 85 7E"),
]


def hdlc_read_firmware_blocking(ip: str, port: int, log_cb):
    """Blocking HDLC read: SNRM → AARQ → RLRQ → NIC Firmware → DISC.
    Returns (ok, message, result_dict|None)."""
    sock = None
    current_frame_name = None
    try:
        sa = resolve_ipv6_sockaddr(ip, port)
        log_cb(f"HDLC  Connecting to [{sa[0]}]:{sa[1]} …")
        sock = socket.socket(socket.AF_INET6, socket.SOCK_STREAM)
        sock.setsockopt(socket.IPPROTO_IPV6, socket.IPV6_V6ONLY, 0)
        sock.settimeout(5)
        sock.connect(sa)
        log_cb("HDLC  Connected ✔")

        result = None
        for name, hexframe in HDLC_FRAMES:
            current_frame_name = name
            frame = bytes.fromhex(hexframe.replace(" ", ""))
            log_cb(f"HDLC  TX [{name}]  {len(frame)} bytes")

            try:
                sock.sendall(frame)
            except OSError as e:
                msg = f"Meter is not sending RX against {name} — TX failed ({type(e).__name__}: {e})."
                log_cb(f"HDLC  ❌ {msg}")
                _safe_close_socket(sock)
                return False, msg, None

            time.sleep(1)

            try:
                rx = sock.recv(4096)
            except socket.timeout:
                rx = b""
            except OSError as e:
                msg = f"Meter is not sending RX against {name} — connection dropped ({type(e).__name__}: {e})."
                log_cb(f"HDLC  ❌ {msg}")
                _safe_close_socket(sock)
                return False, msg, None

            if not rx:
                msg = f"Meter is not sending RX against {name}."
                log_cb(f"HDLC  ❌ {msg}")
                _safe_close_socket(sock)
                return False, msg, None

            log_cb(f"HDLC  RX [{name}]  {len(rx)} bytes  |  HEX: {rx.hex(' ').upper()[:80]}{'…' if len(rx)>40 else ''}")
            if name == "NIC Firmware":
                result = decode_firmware_response(rx)

        _safe_close_socket(sock)
        log_cb("HDLC  Session closed.")
        time.sleep(1.5)

        if result is None or not result.get("firmware"):
            return False, "HDLC sequence completed but no firmware string was returned by the NIC.", result
        return True, "OK", result
    except Exception as exc:
        _safe_close_socket(sock)
        if current_frame_name:
            msg = f"Meter is not sending RX against {current_frame_name} — {type(exc).__name__}: {exc}"
        else:
            msg = f"{type(exc).__name__}: {exc}"
        log_cb(f"HDLC  ❌ {msg}")
        return False, msg, None


def run_nic_firmware_sequence(ip: str, port: int, log_cb, done_cb):
    def _worker():
        result = None
        sock = None
        current_frame_name = None
        try:
            sa = resolve_ipv6_sockaddr(ip, port)
            log_cb(f"NIC-FW  Connecting to [{sa[0]}]:{sa[1]} …")
            sock = socket.socket(socket.AF_INET6, socket.SOCK_STREAM)
            sock.setsockopt(socket.IPPROTO_IPV6, socket.IPV6_V6ONLY, 0)
            sock.settimeout(5)
            sock.connect(sa)
            log_cb("NIC-FW  Connected ✔")

            for name, hexframe in HDLC_FRAMES:
                current_frame_name = name
                frame = bytes.fromhex(hexframe.replace(" ", ""))
                log_cb(f"NIC-FW  TX [{name}]  {len(frame)} bytes")

                try:
                    sock.sendall(frame)
                except OSError as e:
                    msg = f"Meter is not sending RX against {name} — TX failed ({type(e).__name__}: {e})."
                    log_cb(f"NIC-FW  ❌ {msg}")
                    _safe_close_socket(sock)
                    done_cb(None, msg)
                    return

                time.sleep(1)

                try:
                    rx = sock.recv(4096)
                except socket.timeout:
                    rx = b""
                except OSError as e:
                    msg = f"Meter is not sending RX against {name} — connection dropped ({type(e).__name__}: {e})."
                    log_cb(f"NIC-FW  ❌ {msg}")
                    _safe_close_socket(sock)
                    done_cb(None, msg)
                    return

                if not rx:
                    msg = f"Meter is not sending RX against {name}."
                    log_cb(f"NIC-FW  ❌ {msg}")
                    _safe_close_socket(sock)
                    done_cb(None, msg)
                    return

                log_cb(f"NIC-FW  RX [{name}]  {len(rx)} bytes  |  HEX: {rx.hex(' ').upper()[:80]}{'…' if len(rx)>40 else ''}")
                if name == "NIC Firmware":
                    result = decode_firmware_response(rx)

            _safe_close_socket(sock)
            log_cb("NIC-FW  Session closed.")
        except Exception as exc:
            _safe_close_socket(sock)
            if current_frame_name:
                err_msg = f"Meter is not sending RX against {current_frame_name} — {type(exc).__name__}: {exc}"
            else:
                err_msg = f"{type(exc).__name__}: {exc}"
            log_cb(f"NIC-FW  ❌ {err_msg}")
            done_cb(None, err_msg)
            return

        if result is None or not result.get("firmware"):
            done_cb(None, "HDLC sequence completed but no firmware string was returned by the NIC.")
            return
        done_cb(result, None)

    threading.Thread(target=_worker, daemon=True).start()


# ─────────────────────────────────────────────────────────────
#  Styled widget building blocks
# ─────────────────────────────────────────────────────────────
class Card(QtWidgets.QFrame):
    """A consistently-styled rounded 'card' panel used throughout the UI."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"""
            QFrame#card {{
                background-color: {CARD_BG};
                border: 1px solid {CARD_BORDER};
                border-radius: 14px;
            }}
        """)
        self.setObjectName("card")
        shadow = QtWidgets.QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(18)
        shadow.setOffset(0, 3)
        shadow.setColor(QtGui.QColor(0, 0, 0, 28))
        self.setGraphicsEffect(shadow)


class Label(QtWidgets.QLabel):
    def __init__(self, text="", font=None, color=None, textvariable=None, parent=None):
        super().__init__(text, parent)
        self.setFont(font() if font else FONT_BODY())
        self._color = color
        if color:
            self.setStyleSheet(f"color: {color};")
        if textvariable is not None:
            self.setText(textvariable.get())
            textvariable.changed.connect(self.setText)

    def set_color(self, color):
        self._color = color
        self.setStyleSheet(f"color: {color};")


class StatusPill(QtWidgets.QLabel):
    def __init__(self, text="", bg="#8a9c5c", parent=None):
        super().__init__(text, parent)
        self.setFont(FONT_LABEL())
        self.setAlignment(Qt.AlignCenter)
        self._apply(bg)

    def _apply(self, bg):
        self.setStyleSheet(f"""
            QLabel {{
                color: white; background-color: {bg};
                border-radius: 13px; padding: 6px 16px;
            }}
        """)

    def set_status(self, text, bg):
        self.setText(text)
        self._apply(bg)


class Button(QtWidgets.QPushButton):
    def __init__(self, text="", fg_color=BRAND, hover_color=BRAND_HOVER,
                 width=None, height=36, font=None, parent=None):
        super().__init__(text, parent)
        self.setFont(font() if font else FONT_LABEL())
        self.setCursor(Qt.PointingHandCursor)
        self.setFixedHeight(height)
        if width:
            self.setMinimumWidth(width)
        self._fg, self._hover = fg_color, hover_color
        self._apply_style()

    def _apply_style(self):
        self.setStyleSheet(f"""
            QPushButton {{
                background-color: {self._fg}; color: white; border: none;
                border-radius: 9px; padding: 6px 16px; font-weight: 600;
            }}
            QPushButton:hover:!disabled {{ background-color: {self._hover}; }}
            QPushButton:disabled {{ background-color: #cfd3c8; color: #90948a; }}
        """)

    def set_colors(self, fg_color, hover_color):
        self._fg, self._hover = fg_color, hover_color
        self._apply_style()


class LineEdit(QtWidgets.QLineEdit):
    def __init__(self, text="", width=None, mono=False, parent=None):
        super().__init__(text, parent)
        self.setFont(FONT_MONO() if mono else FONT_BODY())
        self.setFixedHeight(34)
        if width:
            self.setFixedWidth(width)
        self.setStyleSheet(f"""
            QLineEdit {{
                border: 1px solid {CARD_BORDER}; border-radius: 8px;
                padding: 4px 10px; background: white; color: {TEXT_DARK};
            }}
            QLineEdit:focus {{ border: 1.5px solid {BRAND}; }}
            QLineEdit:disabled {{ background: #f0f1ec; color: #9a9d94; }}
            QLineEdit:read-only {{ background: #f6f8f4; color: #555; }}
        """)


class ComboBox(QtWidgets.QComboBox):
    def __init__(self, values=None, width=None, parent=None):
        super().__init__(parent)
        self.setFont(FONT_BODY())
        self.setFixedHeight(34)
        if width:
            self.setFixedWidth(width)
        if values:
            self.addItems(values)
        self.setStyleSheet(f"""
            QComboBox {{
                border: 1px solid {CARD_BORDER}; border-radius: 8px;
                padding: 4px 10px; background: white; color: {TEXT_DARK};
            }}
            QComboBox:focus {{ border: 1.5px solid {BRAND}; }}
            QComboBox::drop-down {{ border: none; width: 24px; }}
            QComboBox QAbstractItemView {{
                border: 1px solid {CARD_BORDER}; selection-background-color: {BRAND_LIGHT};
                selection-color: {TEXT_DARK};
            }}
        """)


class TextBox(QtWidgets.QPlainTextEdit):
    def __init__(self, mono=True, height=None, readonly=False, parent=None):
        super().__init__(parent)
        self.setFont(FONT_MONO() if mono else FONT_BODY())
        if height:
            self.setFixedHeight(height)
        self.setReadOnly(readonly)
        self.setStyleSheet(f"""
            QPlainTextEdit {{
                border: 1px solid {CARD_BORDER}; border-radius: 10px;
                padding: 8px; background: #fbfcfa; color: {TEXT_DARK};
            }}
        """)

    def append_line(self, text):
        self.moveCursor(QtGui.QTextCursor.End)
        self.insertPlainText(text)
        self.moveCursor(QtGui.QTextCursor.End)
        self.ensureCursorVisible()

    def clear_all(self):
        self.clear()


class ProgressBar(QtWidgets.QProgressBar):
    def __init__(self, height=14, parent=None):
        super().__init__(parent)
        self.setRange(0, 1000)
        self.setTextVisible(False)
        self.setFixedHeight(height)
        self._color = BRAND
        self._apply()

    def set_fraction(self, frac):
        self.setValue(int(max(0.0, min(1.0, frac)) * 1000))

    def set_color(self, color):
        self._color = color
        self._apply()

    def _apply(self):
        self.setStyleSheet(f"""
            QProgressBar {{
                background: #e2e5df; border-radius: 7px; border: none;
            }}
            QProgressBar::chunk {{
                background-color: {self._color}; border-radius: 7px;
            }}
        """)


# ─────────────────────────────────────────────────────────────
#  NIC Firmware Result Dialog
# ─────────────────────────────────────────────────────────────
class NicFirmwareDialog(QtWidgets.QDialog):
    def __init__(self, parent, result, error_msg=None):
        super().__init__(parent)
        self.setWindowTitle("NIC Firmware Info")
        self.resize(680, 460)
        self.setStyleSheet(f"QDialog {{ background-color: {BG_MAIN}; }}")

        root = QtWidgets.QVBoxLayout(self)
        root.setContentsMargins(24, 20, 24, 18)
        root.setSpacing(10)

        root.addWidget(Label("📟  NIC Firmware Details", font=FONT_H1))

        if result is None:
            root.addSpacing(10)
            err_title = Label("❌  Could not retrieve firmware version.", font=FONT_BODY, color=DANGER)
            root.addWidget(err_title)
            err_box = Label(error_msg or "Unknown error — check logs.", font=FONT_MONO, color=DANGER)
            err_box.setWordWrap(True)
            root.addWidget(err_box)
        else:
            card = Card(self)
            grid = QtWidgets.QGridLayout(card)
            grid.setContentsMargins(16, 12, 16, 12)
            grid.setVerticalSpacing(10)

            def row(r, label, value, color=None):
                grid.addWidget(Label(label, font=FONT_LABEL), r, 0, alignment=Qt.AlignLeft)
                v = Label(value or "—", font=FONT_BODY, color=color)
                v.setWordWrap(True)
                grid.addWidget(v, r, 1, alignment=Qt.AlignLeft)

            fw = result.get("firmware")
            row(0, "🔧  Firmware",     fw or "Not found", SUCCESS if fw else DANGER)
            row(1, "🆔  IMEI",         result.get("imei") or "Not found")
            row(2, "🏭  Manufacturer", result.get("manufacturer") or "Not found")
            root.addWidget(card)

            root.addWidget(Label("ASCII representation:", font=FONT_LABEL))
            ascii_box = TextBox(mono=True, height=70, readonly=True)
            ascii_box.setPlainText(result.get("ascii", ""))
            root.addWidget(ascii_box)

            root.addWidget(Label("Raw HEX:", font=FONT_LABEL))
            hex_box = TextBox(mono=True, height=60, readonly=True)
            hex_box.setPlainText(result.get("raw_hex", ""))
            root.addWidget(hex_box)

        root.addStretch(1)
        btn_row = QtWidgets.QHBoxLayout()
        btn_row.addStretch(1)
        close_btn = Button("Close", fg_color=SLATE, hover_color=SLATE_HOVER, width=110)
        close_btn.clicked.connect(self.accept)
        btn_row.addWidget(close_btn)
        btn_row.addStretch(1)
        root.addLayout(btn_row)


# ─────────────────────────────────────────────────────────────
#  Ping Dialog
# ─────────────────────────────────────────────────────────────
class PingDialog(QtWidgets.QDialog):
    _log_signal = QtCore.Signal(str)
    _status_signal = QtCore.Signal(object)
    _reenable_signal = QtCore.Signal()

    def __init__(self, parent, default_ip=""):
        super().__init__(parent)
        self.setWindowTitle("Ping Tool")
        self.resize(660, 480)
        self.setStyleSheet(f"QDialog {{ background-color: {BG_MAIN}; }}")

        self._log_signal.connect(self._append)
        self._status_signal.connect(self._set_status)
        self._reenable_signal.connect(lambda: self.ping_btn.setEnabled(True))

        root = QtWidgets.QVBoxLayout(self)
        root.setContentsMargins(18, 16, 18, 14)
        root.setSpacing(10)

        root.addWidget(Label("🔔  Ping Tool", font=FONT_H1))

        top = Card(self)
        top_l = QtWidgets.QHBoxLayout(top)
        top_l.setContentsMargins(14, 12, 14, 12)
        top_l.addWidget(Label("Address:", font=FONT_LABEL))
        self.ip_edit = LineEdit(default_ip, width=320)
        top_l.addWidget(self.ip_edit)
        top_l.addWidget(Label("Count:", font=FONT_LABEL))
        self.count_edit = LineEdit("4", width=60)
        top_l.addWidget(self.count_edit)
        self.ping_btn = Button("📡  Ping", fg_color=INFO, hover_color=INFO_HOVER, width=100)
        self.ping_btn.clicked.connect(self.run_ping)
        top_l.addWidget(self.ping_btn)
        top_l.addStretch(1)
        root.addWidget(top)

        self.status_label = Label("", font=FONT_H2)
        root.addWidget(self.status_label)

        out_card = Card(self)
        out_l = QtWidgets.QVBoxLayout(out_card)
        out_l.setContentsMargins(12, 10, 12, 10)
        out_l.addWidget(Label("Output", font=FONT_LABEL))
        self.output_box = TextBox(mono=True, readonly=True)
        out_l.addWidget(self.output_box)
        root.addWidget(out_card, 1)

        btn_row = QtWidgets.QHBoxLayout()
        clear_btn = Button("Clear", fg_color=SLATE, hover_color=SLATE_HOVER, width=90)
        clear_btn.clicked.connect(self.output_box.clear_all)
        btn_row.addWidget(clear_btn)
        btn_row.addStretch(1)
        close_btn = Button("Close", fg_color=SLATE, hover_color=SLATE_HOVER, width=90)
        close_btn.clicked.connect(self.accept)
        btn_row.addWidget(close_btn)
        root.addLayout(btn_row)

    def _append(self, text):
        self.output_box.append_line(text)

    def _set_status(self, success):
        if success is True:
            self.status_label.setText("✅  Host reachable")
            self.status_label.set_color(SUCCESS)
        elif success is False:
            self.status_label.setText("❌  Host unreachable / timeout")
            self.status_label.set_color(DANGER)
        else:
            self.status_label.setText("⏳  Pinging…")
            self.status_label.set_color(WARNING)

    def run_ping(self):
        ip = self.ip_edit.text().strip()
        if not ip:
            self._append("⚠  No address entered.\n")
            return
        try:
            count = int(self.count_edit.text().strip())
        except ValueError:
            count = 4
        self.output_box.clear_all()
        self._set_status(None)
        self.ping_btn.setEnabled(False)
        threading.Thread(target=self._ping_worker, args=(ip, count), daemon=True).start()

    def _ping_worker(self, ip, count):
        is_win = platform.system().lower() == "windows"
        is_v6  = ":" in ip
        if is_v6:
            if is_win:
                candidates = [["ping", "-6", "-n", str(count), ip]]
            else:
                candidates = [["ping6", "-c", str(count), ip],
                              ["ping",  "-6", "-c", str(count), ip]]
        else:
            flag = "-n" if is_win else "-c"
            candidates = [["ping", flag, str(count), ip]]

        proc = None
        for cmd in candidates:
            self._log_signal.emit("$ " + " ".join(cmd) + "\n\n")
            try:
                proc = subprocess.Popen(cmd, stdout=subprocess.PIPE,
                                        stderr=subprocess.STDOUT, text=True)
                break
            except FileNotFoundError:
                self._log_signal.emit("Not found: " + cmd[0] + "\n")
                proc = None

        if proc is None:
            self._log_signal.emit("⚠  No suitable ping command found.\n")
            self._status_signal.emit(False)
            self._reenable_signal.emit()
            return

        success = None
        for line in proc.stdout:
            self._log_signal.emit(line)
            ll = line.lower()
            if any(k in ll for k in ("ttl=", "time=", "bytes from")):
                success = True
            if any(k in ll for k in ("unreachable", "100% packet loss",
                                      "request timed out", "unknown host",
                                      "network is unreachable")):
                success = False
        proc.wait()
        if success is None:
            success = (proc.returncode == 0)
        self._status_signal.emit(success)
        self._reenable_signal.emit()


# ─────────────────────────────────────────────────────────────
#  Stress Test — HDLC frame builder (CRC-verified against real captures)
# ─────────────────────────────────────────────────────────────
def _fcs16_table():
    poly = 0x8408
    table = []
    for byte in range(256):
        crc = byte
        for _ in range(8):
            crc = (crc >> 1) ^ poly if (crc & 1) else (crc >> 1)
        table.append(crc & 0xFFFF)
    return table


_FCS_TABLE = _fcs16_table()


def fcs16(data: bytes) -> int:
    crc = 0xFFFF
    for b in data:
        crc = (crc >> 8) ^ _FCS_TABLE[(crc ^ b) & 0xFF]
    return (~crc) & 0xFFFF


def _le(crc: int) -> bytes:
    return bytes([crc & 0xFF, (crc >> 8) & 0xFF])


# Address field used by this client for every TX frame (fixed for this NIC).
STRESS_ADDR_TX = bytes.fromhex("0002 0401 FD".replace(" ", ""))

# Info field (LLC + xDLMS APDU) of the "NIC Firmware" GET-request —
# everything between HCS and FCS in the known-good captured frame.
STRESS_NIC_FW_INFO = bytes.fromhex("E6E600C001C100010087640000FF0200")


class HDLCSession:
    """Tracks HDLC N(S)/N(R) sequence numbers and builds correctly
    checksummed I-frames — required because resending an identical frame
    (same sequence numbers) makes the NIC treat it as a duplicate and
    stop responding after the first one."""

    def __init__(self, addr_tx: bytes):
        self.addr_tx = addr_tx
        self.ns = 0
        self.nr = 0

    def build_iframe(self, info: bytes, poll: int = 1) -> bytes:
        control = ((self.nr & 0x7) << 5) | ((poll & 0x1) << 4) | ((self.ns & 0x7) << 1)
        addr_len = len(self.addr_tx)
        ll = addr_len + 7 + len(info)
        header = bytes([0xA0, ll & 0xFF]) + self.addr_tx + bytes([control])
        hcs = _le(fcs16(header))
        body = header + hcs + info
        fcs = _le(fcs16(body))
        return b"\x7e" + body + fcs + b"\x7e"

    def advance_after_send(self):
        self.ns = (self.ns + 1) % 8

    def update_from_response(self, resp: bytes, addr_len: int = 5):
        if len(resp) < 4 or resp[0] != 0x7E:
            return None
        inner = resp[1:-1] if resp[-1] == 0x7E else resp[1:]
        ctrl_index = 2 + addr_len
        if len(inner) <= ctrl_index:
            return None
        control = inner[ctrl_index]
        if control & 0x1:
            return control  # S-/U-frame, no N(S) to ack
        server_ns = (control >> 1) & 0x7
        self.nr = (server_ns + 1) % 8
        return control


# ─────────────────────────────────────────────────────────────
#  Stress Test Dialog
# ─────────────────────────────────────────────────────────────
class StressTestDialog(QtWidgets.QDialog):
    _log_signal = QtCore.Signal(str)
    _stats_signal = QtCore.Signal(dict)
    _finished_signal = QtCore.Signal(bool)   # True = completed normally, False = aborted on error

    UNIT_SECONDS = {"Seconds": 1, "Minutes": 60, "Hours": 3600}

    def __init__(self, parent, ip="", port=4059):
        super().__init__(parent)
        self.setWindowTitle("Stress Test")
        self.resize(760, 640)
        self.setStyleSheet(f"QDialog {{ background-color: {BG_MAIN}; }}")

        self._cancel_flag = False
        self._running = False
        self._attempts = 0
        self._success = 0
        self._failed = 0
        self._response_times = []

        self._log_signal.connect(self._append_log)
        self._stats_signal.connect(self._update_stats)
        self._finished_signal.connect(self._on_finished)

        root = QtWidgets.QVBoxLayout(self)
        root.setContentsMargins(20, 18, 20, 16)
        root.setSpacing(12)

        root.addWidget(Label("🔥  NIC Stress Lab", font=FONT_H1))
        root.addWidget(Label("Hammer the NIC with repeated firmware-version requests and see how it holds up.",
                              font=FONT_SMALL, color=MUTED_TEXT))

        # ── Config card ──
        cfg = Card(self)
        cfg_l = QtWidgets.QGridLayout(cfg)
        cfg_l.setContentsMargins(16, 14, 16, 14)
        cfg_l.setVerticalSpacing(10)
        cfg_l.setHorizontalSpacing(10)

        cfg_l.addWidget(Label("🌐  Target", font=FONT_LABEL), 0, 0)
        self.target_label = Label(f"[{ip}]:{port}", font=FONT_MONO, color=TEXT_DARK)
        cfg_l.addWidget(self.target_label, 0, 1, 1, 3)

        cfg_l.addWidget(Label("⏱️  Run for", font=FONT_LABEL), 1, 0)
        self.duration_spin = QtWidgets.QSpinBox()
        self.duration_spin.setRange(1, 999999)
        self.duration_spin.setValue(60)
        self.duration_spin.setFixedHeight(34)
        self.duration_spin.setFixedWidth(100)
        self.duration_spin.setStyleSheet(f"""
            QSpinBox {{
                border: 1px solid {CARD_BORDER}; border-radius: 8px;
                padding: 4px 8px; background: white; color: {TEXT_DARK};
            }}
            QSpinBox:focus {{ border: 1.5px solid {BRAND}; }}
        """)
        cfg_l.addWidget(self.duration_spin, 1, 1)

        self.unit_combo = ComboBox(values=list(self.UNIT_SECONDS.keys()), width=120)
        cfg_l.addWidget(self.unit_combo, 1, 2)

        self.duration_hint = Label("= 60 sec", font=FONT_SMALL, color=MUTED_TEXT)
        cfg_l.addWidget(self.duration_hint, 1, 3)
        self.duration_spin.valueChanged.connect(self._update_duration_hint)
        self.unit_combo.currentTextChanged.connect(self._update_duration_hint)

        cfg_l.setColumnStretch(3, 1)
        root.addWidget(cfg)

        # ── Stats card ──
        stats = Card(self)
        st_l = QtWidgets.QGridLayout(stats)
        st_l.setContentsMargins(16, 12, 16, 12)
        st_l.setHorizontalSpacing(22)

        def stat(col, label):
            box = QtWidgets.QVBoxLayout()
            box.setSpacing(2)
            l = Label(label, font=FONT_SMALL, color=MUTED_TEXT)
            v = Label("—", font=FONT_H2, color=TEXT_DARK)
            box.addWidget(l)
            box.addWidget(v)
            st_l.addLayout(box, 0, col)
            return v

        self.stat_attempts = stat(0, "🎯  Attempts")
        self.stat_success  = stat(1, "✅  Success")
        self.stat_failed   = stat(2, "❌  Failed")
        self.stat_rate     = stat(3, "📊  Success Rate")
        self.stat_avg      = stat(4, "⏳  Avg ms")
        self.stat_minmax   = stat(5, "↕️  Min / Max ms")
        st_l.setColumnStretch(6, 1)
        root.addWidget(stats)

        # ── Buttons ──
        btn_row = QtWidgets.QHBoxLayout()
        self.start_btn = Button("▶️  Start", fg_color=BRAND, hover_color=BRAND_HOVER, width=110)
        self.start_btn.clicked.connect(self.start_test)
        btn_row.addWidget(self.start_btn)

        self.stop_btn = Button("⏹  Stop", fg_color=DANGER, hover_color=DANGER_HOVER, width=100)
        self.stop_btn.clicked.connect(self.stop_test)
        self.stop_btn.setEnabled(False)
        btn_row.addWidget(self.stop_btn)

        btn_row.addStretch(1)

        self.copy_btn = Button("📋  Copy Logs", fg_color=INFO, hover_color=INFO_HOVER, width=120)
        self.copy_btn.clicked.connect(self.copy_logs)
        btn_row.addWidget(self.copy_btn)

        self.clear_btn = Button("Clear", fg_color=SLATE, hover_color=SLATE_HOVER, width=90)
        self.clear_btn.clicked.connect(lambda: self.log_box.clear_all())
        btn_row.addWidget(self.clear_btn)

        root.addLayout(btn_row)

        # ── Log ──
        log_card = Card(self)
        log_l = QtWidgets.QVBoxLayout(log_card)
        log_l.setContentsMargins(12, 10, 12, 10)
        log_l.addWidget(Label("📝  Log  (selectable — Ctrl+A, Ctrl+C, or use Copy Logs)", font=FONT_LABEL))
        self.log_box = TextBox(mono=True, readonly=True)
        self.log_box.setLineWrapMode(QtWidgets.QPlainTextEdit.NoWrap)
        log_l.addWidget(self.log_box)
        root.addWidget(log_card, 1)

        close_row = QtWidgets.QHBoxLayout()
        close_row.addStretch(1)
        self.close_btn = Button("Close", fg_color=SLATE, hover_color=SLATE_HOVER, width=100)
        self.close_btn.clicked.connect(self.reject)
        close_row.addWidget(self.close_btn)
        root.addLayout(close_row)

        self._update_duration_hint()

    # ── UI helpers ──
    def _update_duration_hint(self, *_):
        secs = self.duration_spin.value() * self.UNIT_SECONDS[self.unit_combo.currentText()]
        self.duration_hint.setText(f"= {secs:,} sec")

    def _append_log(self, text):
        self.log_box.append_line(text)

    def _update_stats(self, d):
        self.stat_attempts.setText(str(d["attempts"]))
        self.stat_success.setText(str(d["success"]))
        self.stat_failed.setText(str(d["failed"]))
        rate = (d["success"] / d["attempts"] * 100) if d["attempts"] else 0.0
        self.stat_rate.setText(f"{rate:.1f}%")
        self.stat_rate.set_color(SUCCESS if rate >= 90 else (WARNING if rate >= 50 else DANGER))
        rt = d["response_times"]
        if rt:
            self.stat_avg.setText(f"{sum(rt)/len(rt):.1f}")
            self.stat_minmax.setText(f"{min(rt):.1f} / {max(rt):.1f}")
        else:
            self.stat_avg.setText("—")
            self.stat_minmax.setText("—")

    def copy_logs(self):
        QtWidgets.QApplication.clipboard().setText(self.log_box.toPlainText())
        self._append_log("📋  Logs copied to clipboard.\n")

    def reject(self):
        if self._running:
            if not mb.askyesno("Stress test running",
                                "A stress test is currently running.\n\nClose anyway? "
                                "(it will keep running until it finishes or is stopped)"):
                return
        super().reject()

    # ── Control ──
    def start_test(self):
        if self._running:
            return
        ip, port = self._resolve_target()
        if not ip:
            self._append_log("⚠  No NIC IPv6 address available from the main form.\n")
            return

        seconds = self.duration_spin.value() * self.UNIT_SECONDS[self.unit_combo.currentText()]

        self._cancel_flag = False
        self._running = True
        self._attempts = 0
        self._success = 0
        self._failed = 0
        self._response_times = []
        self._update_stats(dict(attempts=0, success=0, failed=0, response_times=[]))

        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.duration_spin.setEnabled(False)
        self.unit_combo.setEnabled(False)

        self._append_log("=" * 60 + "\n")
        self._append_log(f"🔥  Starting stress test against [{ip}]:{port} for {seconds:,} sec ...\n")

        threading.Thread(target=self._worker, args=(ip, port, seconds), daemon=True).start()

    def stop_test(self):
        if not self._running:
            return
        self._cancel_flag = True
        self.stop_btn.setEnabled(False)
        self._log_signal.emit("⏹  Stop requested — finishing current request …\n")

    def _resolve_target(self):
        parent = self.parent()
        if parent is not None and hasattr(parent, "_get_ip_port"):
            return parent._get_ip_port()
        return "", 4059

    # ── Worker (background thread) ──
# ── UI helpers ──
    @staticmethod
    def _ts():
        return datetime.now().strftime("%H:%M:%S.%f")[:-3]  # HH:MM:SS.mmm

    def _append_log(self, text):
        self.log_box.append_line(text)

    def _log_ts(self, text):
        """Emit a single log line prefixed with a millisecond timestamp."""
        self._log_signal.emit(f"[{self._ts()}] {text}")

    # ── Worker (background thread) ──
    def _worker(self, ip, port, duration_seconds):
        sock = None
        try:
            sa = resolve_ipv6_sockaddr(ip, port)
            self._log_ts(f"Connecting to [{sa[0]}]:{sa[1]} …\n")
            sock = make_tcp_socket(timeout=5)
            sock.connect(sa)
            self._log_ts("Connected ✔\n")

            session = HDLCSession(STRESS_ADDR_TX)

            handshake = [f for f in HDLC_FRAMES if f[0] in ("SNRM", "AARQ", "RLRQ")]
            for name, hexframe in handshake:
                frame = bytes.fromhex(hexframe.replace(" ", ""))
                self._log_ts(f"TX [{name}]  {len(frame)}B\n")
                self._log_ts(f"    → {frame.hex(' ').upper()}\n")

                sock.sendall(frame)
                rx = self._recv_frame(sock)

                if not rx:
                    self._log_ts(f"❌  No response to {name} — aborting.\n")
                    _safe_close_socket(sock)
                    self._finished_signal.emit(False)
                    return

                self._log_ts(f"RX [{name}]  {len(rx)}B\n")
                self._log_ts(f"    ← {rx.hex(' ').upper()}\n")
                self._log_ts(f"    ASCII: {printable_ascii(rx)}\n")

                session.update_from_response(rx)
                if name != "SNRM":
                    session.advance_after_send()

            self._log_ts("Handshake complete. Beginning firmware-request loop …\n")
            self._append_log("-" * 60 + "\n")

            test_start = time.time()
            while time.time() - test_start < duration_seconds:
                if self._cancel_flag:
                    self._log_ts("⏹  Stopped by user.\n")
                    break

                self._attempts += 1
                frame = session.build_iframe(STRESS_NIC_FW_INFO, poll=1)

                self._log_ts(f"TX [#{self._attempts}]  {len(frame)}B\n")
                self._log_ts(f"    → {frame.hex(' ').upper()}\n")

                t0 = time.perf_counter()
                sock.sendall(frame)
                rx = self._recv_frame(sock)
                elapsed_ms = (time.perf_counter() - t0) * 1000

                if rx:
                    self._success += 1
                    self._response_times.append(elapsed_ms)
                    session.advance_after_send()
                    session.update_from_response(rx)

                    self._log_ts(f"RX [#{self._attempts}]  {len(rx)}B  ({elapsed_ms:.1f} ms)\n")
                    self._log_ts(f"    ← {rx.hex(' ').upper()}\n")
                    self._log_ts(f"    ASCII: {printable_ascii(rx)}\n")
                else:
                    self._failed += 1
                    self._log_ts(f"RX [#{self._attempts}]  ❌  no response  (timeout after {elapsed_ms:.1f} ms)\n")

                self._append_log("\n")  # blank line between attempts for readability

                self._stats_signal.emit(dict(
                    attempts=self._attempts, success=self._success,
                    failed=self._failed, response_times=list(self._response_times)))

                time.sleep(0.2)

            # Best-effort graceful close
            try:
                disc = bytes.fromhex(next(f[1] for f in HDLC_FRAMES if f[0] == "DISC").replace(" ", ""))
                self._log_ts(f"TX [DISC]  {len(disc)}B\n")
                self._log_ts(f"    → {disc.hex(' ').upper()}\n")
                sock.sendall(disc)
                rx = self._recv_frame(sock)
                if rx:
                    self._log_ts(f"RX [DISC]  {len(rx)}B\n")
                    self._log_ts(f"    ← {rx.hex(' ').upper()}\n")
            except Exception:
                pass

            _safe_close_socket(sock)

            self._append_log("-" * 60 + "\n")
            self._append_log("Stress Test Summary\n")
            self._append_log(f"  Attempts       : {self._attempts}\n")
            self._append_log(f"  Successful     : {self._success}\n")
            self._append_log(f"  Failed         : {self._failed}\n")
            if self._attempts:
                self._append_log(f"  Success Rate   : {self._success/self._attempts*100:.2f}%\n")
            if self._response_times:
                rt = self._response_times
                self._append_log(f"  Avg Response   : {sum(rt)/len(rt):.2f} ms\n")
                self._append_log(f"  Min Response   : {min(rt):.2f} ms\n")
                self._append_log(f"  Max Response   : {max(rt):.2f} ms\n")
            self._append_log("=" * 60 + "\n")

            self._finished_signal.emit(True)

        except Exception as e:
            _safe_close_socket(sock)
            self._log_ts(f"❌  {type(e).__name__}: {e}\n")
            self._finished_signal.emit(False)

    def _recv_frame(self, sock, timeout=5):
        data = bytearray()
        start = time.time()
        sock.settimeout(0.2)
        while True:
            try:
                chunk = sock.recv(4096)
                if chunk:
                    data.extend(chunk)
                    if data[-1] == 0x7E and len(data) > 1:
                        break
            except socket.timeout:
                pass
            except OSError:
                break
            if time.time() - start > timeout:
                break
        return bytes(data)

    def _on_finished(self, ok):
        self._running = False
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.duration_spin.setEnabled(True)
        self.unit_combo.setEnabled(True)


# ─────────────────────────────────────────────────────────────
#  Bulk FOTA Dialog — Excel IPv6 list → per-device FOTA rollout
# ─────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────
#  Bulk FOTA Dialog — Excel IPv6 list → parallel per-device FOTA rollout
# ─────────────────────────────────────────────────────────────
class BulkFotaDialog(QtWidgets.QDialog):
    _log_signal   = QtCore.Signal(str)
    _cell_signal  = QtCore.Signal(int, int, str, str)   # row, col, text, color ("" = default)
    _finished_signal = QtCore.Signal()

    COL_IPV6, COL_CURRENT_FW, COL_STATUS, COL_LATEST_FW = range(4)
    MAX_PARALLEL = 20   # concurrency cap so we don't open hundreds of sockets at once

    def __init__(self, parent, default_ip_port=("", 4059)):
        super().__init__(parent)
        self.setWindowTitle("Bulk FOTA")
        self.resize(680, 680)
        self.setStyleSheet(f"QDialog {{ background-color: {BG_MAIN}; }}")

        self._cancel_flag = False
        self._running = False
        self.ip_list = []
        self._active_count = 0
        self._active_lock = threading.Lock()

        self._log_signal.connect(self._append_log)
        self._cell_signal.connect(self._set_cell)
        self._finished_signal.connect(self._on_finished)

        # ── Outer layout: just the scroll area, edge to edge ──
        outer = QtWidgets.QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QtWidgets.QFrame.NoFrame)
        scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        outer.addWidget(scroll)

        content = QtWidgets.QWidget()
        scroll.setWidget(content)

        root = QtWidgets.QVBoxLayout(content)
        root.setContentsMargins(14, 12, 14, 10)
        root.setSpacing(6)                       # ← tighter vertical rhythm throughout

        root.addWidget(Label("🛰️  Bulk FOTA Rollout", font=FONT_H2))
        root.addWidget(Label("Upload a list of NIC IPv6 addresses — all devices are updated in parallel.",
                              font=FONT_SMALL, color=MUTED_TEXT))

        # ── Excel upload ──
        up_card = Card(self)
        up_l = QtWidgets.QHBoxLayout(up_card)
        up_l.setContentsMargins(8, 6, 8, 6)
        up_l.setSpacing(6)
        up_l.addWidget(Label("📁  IPv6 List (.xlsx)", font=FONT_LABEL))
        self.upload_btn = Button("Upload Excel", fg_color=INFO, hover_color=INFO_HOVER, width=140)
        self.upload_btn.clicked.connect(self.upload_excel)
        up_l.addWidget(self.upload_btn)
        self.upload_label = Label("No file loaded", font=FONT_SMALL, color=MUTED_TEXT)
        up_l.addWidget(self.upload_label)
        up_l.addStretch(1)
        up_l.addWidget(Label("NIC Port", font=FONT_LABEL))
        self.nic_port_entry = LineEdit(str(default_ip_port[1]), width=80)
        up_l.addWidget(self.nic_port_entry)
        root.addWidget(up_card)

        # ── FOTA command parameters ──
        cmd_card = Card(self)
        cmd_l = QtWidgets.QGridLayout(cmd_card)
        cmd_l.setContentsMargins(8, 6, 8, 6)
        cmd_l.setVerticalSpacing(4)               # ← tighter
        cmd_l.setHorizontalSpacing(6)
        cmd_l.addWidget(Label("🛰️  FOTA Command Parameters", font=FONT_LABEL), 0, 0, 1, 4)

        self.entries = {}
        left_col = [("User", "0"), ("Password", "0"), ("Server IP", "0"), ("File Port", "0")]
        for row, (label, default) in enumerate(left_col, start=1):
            cmd_l.addWidget(Label(label, font=FONT_SMALL), row, 0)
            entry = LineEdit(default)
            entry.textChanged.connect(lambda _t: self.update_preview())
            cmd_l.addWidget(entry, row, 1)
            self.entries[label] = entry

        cmd_l.addWidget(Label("FOTA Mode", font=FONT_SMALL), 4, 2)
        self.mode_combo = ComboBox(values=["0 - Full FOTA", "1 - Delta FOTA"])
        self.mode_combo.currentTextChanged.connect(lambda _t: self.update_preview())
        cmd_l.addWidget(self.mode_combo, 4, 3)

        cmd_l.addWidget(Label("File URL", font=FONT_SMALL), 1, 2)
        self.fp_url_entry = LineEdit("https://nicfota.pythonanywhere.com")
        self.fp_url_entry.textChanged.connect(lambda _t: self.update_preview())
        cmd_l.addWidget(self.fp_url_entry, 1, 3)

        cmd_l.addWidget(Label("Function", font=FONT_SMALL), 2, 2)
        self.fp_func_entry = LineEdit("getFile")
        self.fp_func_entry.textChanged.connect(lambda _t: self.update_preview())
        cmd_l.addWidget(self.fp_func_entry, 2, 3)

        cmd_l.addWidget(Label("File Name", font=FONT_SMALL), 3, 2)
        self.fp_fname_entry = LineEdit("DFOTA_131_TO_132.pack")
        self.fp_fname_entry.textChanged.connect(lambda _t: self.update_preview())
        cmd_l.addWidget(self.fp_fname_entry, 3, 3)

        cmd_l.setColumnStretch(1, 1)
        cmd_l.setColumnStretch(3, 1)
        root.addWidget(cmd_card)

        prev_card = Card(self)
        prev_l = QtWidgets.QVBoxLayout(prev_card)
        prev_l.setContentsMargins(8, 5, 8, 5)
        prev_l.setSpacing(3)
        prev_l.addWidget(Label("👁️  Command Preview", font=FONT_LABEL))
        self.command_preview = LineEdit(mono=True)
        self.command_preview.setReadOnly(True)
        prev_l.addWidget(self.command_preview)
        root.addWidget(prev_card)

        # ── Task table ──
        table_card = Card(self)
        table_l = QtWidgets.QVBoxLayout(table_card)
        table_l.setContentsMargins(8, 6, 8, 6)
        table_l.setSpacing(4)
        table_l.addWidget(Label("📋  Devices", font=FONT_LABEL))
        self.table = QtWidgets.QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["IPv6", "Current FW", "FOTA Status", "Latest Firmware"])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.verticalHeader().setDefaultSectionSize(24)   # tighter rows
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.table.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        self.table.setMinimumHeight(90)
        self.table.setStyleSheet(f"""
            QTableWidget {{ border: 1px solid {CARD_BORDER}; border-radius: 8px;
                             background: white; gridline-color: {CARD_BORDER}; }}
            QHeaderView::section {{ background: {BG_MAIN}; padding: 4px; border: none;
                                     font-weight: 600; color: {TEXT_DARK}; }}
        """)
        table_l.addWidget(self.table)
        root.addWidget(table_card)

        # ── Log ──
        log_card = Card(self)
        log_l = QtWidgets.QVBoxLayout(log_card)
        log_l.setContentsMargins(8, 6, 8, 6)
        log_l.setSpacing(3)
        log_l.addWidget(Label("📝  Log", font=FONT_LABEL))
        self.log_box = TextBox(mono=True, height=90, readonly=True)
        log_l.addWidget(self.log_box)
        root.addWidget(log_card)

        # ── Buttons (pinned outside the scroll area, always visible) ──
        btn_bar = QtWidgets.QWidget()
        btn_row = QtWidgets.QHBoxLayout(btn_bar)
        btn_row.setContentsMargins(14, 8, 14, 10)
        self.run_btn = Button("🚀  Run Bulk FOTA", fg_color=BRAND, hover_color=BRAND_HOVER, width=170)
        self.run_btn.clicked.connect(self.start_bulk)
        btn_row.addWidget(self.run_btn)
        self.read_fw_btn = Button("🔍  Read Current FW", fg_color=INFO, hover_color=INFO_HOVER, width=170)
        self.read_fw_btn.clicked.connect(self.start_read_fw)
        btn_row.addWidget(self.read_fw_btn)
        self.stop_btn = Button("⏹  Stop", fg_color=DANGER, hover_color=DANGER_HOVER, width=100)
        self.stop_btn.clicked.connect(self.stop_bulk)
        self.stop_btn.setEnabled(False)
        btn_row.addWidget(self.stop_btn)
        btn_row.addStretch(1)
        close_btn = Button("Close", fg_color=SLATE, hover_color=SLATE_HOVER, width=100)
        close_btn.clicked.connect(self.reject)
        btn_row.addWidget(close_btn)
        outer.addWidget(btn_bar)

        self.update_preview()

    # ── Command preview (self-contained, mirrors main window's logic) ──
    def update_preview(self):
        mode_val = self.mode_combo.currentText().split(" ")[0]
        fname = self.fp_fname_entry.text().strip()
        url = self.fp_url_entry.text().strip().rstrip("/")
        func = self.fp_func_entry.text().strip().strip("/")
        file_path = url + "/" + func + "/" + fname if func else url + "/" + fname
        cmd = (
            "*GPRS-FOTA," + self.entries["User"].text() + ","
            + self.entries["Password"].text() + ","
            + self.entries["Server IP"].text() + ","
            + mode_val + "," + file_path + ","
            + self.entries["File Port"].text() + ",#"
        )
        self.command_preview.setText(cmd)
        return cmd

    # ── Read-only firmware check (no FOTA command sent) ──
    def start_read_fw(self):
        if self._running:
            return
        if not self.ip_list:
            mb.showerror("No devices", "Upload an Excel file with NIC IPv6 addresses first.")
            return
        try:
            port = int(self.nic_port_entry.text().strip() or 4059)
        except ValueError:
            mb.showerror("Invalid port", "NIC Port must be a number.")
            return

        self._cancel_flag = False
        self._running = True
        self._active_count = len(self.ip_list)
        self._finish_message = "Firmware read complete — all devices finished."

        self.run_btn.setEnabled(False)
        self.read_fw_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.upload_btn.setEnabled(False)

        self._append_log("=" * 60 + "\n")
        self._append_log(f"🔍  Reading current firmware from {len(self.ip_list)} device(s) IN PARALLEL "
                          f"(max {self.MAX_PARALLEL} concurrent)\n")

        for row in range(len(self.ip_list)):
            self._cell_signal.emit(row, self.COL_STATUS, "⏳ Queued …", "")

        sema = threading.Semaphore(min(self.MAX_PARALLEL, len(self.ip_list)))
        for row, ip in enumerate(self.ip_list):
            threading.Thread(target=self._run_device_read_only, args=(row, ip, port, sema), daemon=True).start()
            time.sleep(0.15)   # stagger, same reasoning as the bulk FOTA start

    def _run_device_read_only(self, row, ip, port, sema):
        with sema:
            try:
                self._process_read_only(row, ip, port)
            finally:
                self._device_finished()

    def _process_read_only(self, row, ip, port):
        if self._cancel_flag:
            self._cell_signal.emit(row, self.COL_STATUS, "⏹ Skipped (stopped)", WARNING)
            return

        def log_cb(msg):
            self._log_signal.emit(f"[{ip}] {msg}\n")

        self._cell_signal.emit(row, self.COL_STATUS, "⏳ Reading current FW …", "")
        ok, msg, result = self._run_with_retries(
            "Read current firmware",
            lambda: hdlc_read_firmware_blocking(ip, port, log_cb),
            log_cb)

        if not ok:
            self._cell_signal.emit(row, self.COL_CURRENT_FW, "Unreachable", DANGER)
            self._cell_signal.emit(row, self.COL_STATUS, f"❌ Unreachable after 3 attempts: {msg}", DANGER)
            return

        current_fw = result.get("firmware") or "Unknown"
        self._cell_signal.emit(row, self.COL_CURRENT_FW, current_fw, "")
        self._cell_signal.emit(row, self.COL_STATUS, "✅ Read OK", SUCCESS)

    # ── Excel upload ──
    def upload_excel(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Select IPv6 list", "", "Excel Files (*.xlsx *.xls)")
        if not path:
            return
        try:
            from openpyxl import load_workbook
        except ImportError:
            mb.showerror("Missing dependency",
                          "Reading Excel files needs the 'openpyxl' package.\n\n"
                          "Install it with:  pip install openpyxl")
            return
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            ips = []
            for (cell,) in ws.iter_rows(min_col=1, max_col=1, values_only=False):
                val = str(cell.value).strip() if cell.value is not None else ""
                if val and ":" in val:      # skip blanks and any header row like "IPv6"
                    ips.append(val)
            wb.close()
        except Exception as e:
            mb.showerror("Could not read Excel file", f"{type(e).__name__}: {e}")
            return

        seen = set()
        self.ip_list = [ip for ip in ips if not (ip in seen or seen.add(ip))]

        if not self.ip_list:
            mb.showerror("No IPv6 addresses found",
                          "The first column didn't contain anything that looks like an IPv6 address.")
            return

        self.upload_label.setText(f"{len(self.ip_list)} device(s) loaded — {os.path.basename(path)}")
        self._populate_table()

    def _populate_table(self):
        self.table.setRowCount(len(self.ip_list))
        for row, ip in enumerate(self.ip_list):
            self.table.setItem(row, self.COL_IPV6, QtWidgets.QTableWidgetItem(ip))
            for col in (self.COL_CURRENT_FW, self.COL_STATUS, self.COL_LATEST_FW):
                self.table.setItem(row, col, QtWidgets.QTableWidgetItem("—"))

    # ── UI helpers ──
    def _append_log(self, text):
        self.log_box.append_line(text)

    def _set_cell(self, row, col, text, color):
        item = QtWidgets.QTableWidgetItem(text)
        if color:
            item.setForeground(QtGui.QColor(color))
        self.table.setItem(row, col, item)

    def reject(self):
        if self._running:
            if not mb.askyesno("Bulk FOTA running",
                                "A bulk FOTA rollout is currently running.\n\nClose anyway?"):
                return
        super().reject()

    # ── Control ──
# ── Control ──
    def start_bulk(self):
        if self._running:
            return
        if not self.ip_list:
            mb.showerror("No devices", "Upload an Excel file with NIC IPv6 addresses first.")
            return
        try:
            port = int(self.nic_port_entry.text().strip() or 4059)
        except ValueError:
            mb.showerror("Invalid port", "NIC Port must be a number.")
            return

        cmd = self.update_preview()

        # Validate up front so a bad character doesn't silently kill every device thread
        try:
            payload = (cmd + "\r\n").encode("ascii")
        except UnicodeEncodeError as e:
            mb.showerror("Invalid command",
                          f"The FOTA command contains a non-ASCII character and can't be sent over TCP:\n\n{e}\n\n"
                          f"Check the File URL / Function / File Name fields.")
            return

        self._cancel_flag = False
        self._running = True
        self._active_count = len(self.ip_list)

        self.run_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.upload_btn.setEnabled(False)

        self._append_log("=" * 60 + "\n")
        self._append_log(f"🚀  Starting bulk FOTA for {len(self.ip_list)} device(s) IN PARALLEL "
                          f"(max {self.MAX_PARALLEL} concurrent)\n")
        self._append_log(f"    CMD → {cmd}\n")
        self._append_log(f"    TCP payload ({len(payload)} bytes) → {payload!r}\n")

        for row in range(len(self.ip_list)):
            self._cell_signal.emit(row, self.COL_STATUS, "⏳ Queued …", "")

        sema = threading.Semaphore(min(self.MAX_PARALLEL, len(self.ip_list)))
        for row, ip in enumerate(self.ip_list):
            threading.Thread(target=self._run_device, args=(row, ip, port, payload, sema), daemon=True).start()
            time.sleep(0.15)   # small stagger so connections ramp up instead of all firing at once
    
    def stop_bulk(self):
        self._cancel_flag = True
        self.stop_btn.setEnabled(False)
        self._log_signal.emit("⏹  Stop requested — in-flight devices will finish their current step, "
                               "queued devices will be skipped …\n")

    # ── Per-device worker — runs independently, one thread per IP ──
    def _run_device(self, row, ip, port, payload, sema):
        with sema:
            try:
                self._process_one_device(row, ip, port, payload)
            finally:
                self._device_finished()

    def _process_one_device(self, row, ip, port, payload):
        if self._cancel_flag:
            self._cell_signal.emit(row, self.COL_STATUS, "⏹ Skipped (stopped)", WARNING)
            return

        def log_cb(msg):
            self._log_signal.emit(f"[{ip}] {msg}\n")

        # Step 1: read current firmware — this is where an unreachable NIC gets caught
        self._cell_signal.emit(row, self.COL_STATUS, "⏳ Reading current FW …", "")
        ok, msg, result = hdlc_read_firmware_blocking(ip, port, log_cb)
        if not ok:
            self._cell_signal.emit(row, self.COL_CURRENT_FW, "Unreachable", DANGER)
            self._cell_signal.emit(row, self.COL_STATUS, f"❌ Unreachable — stopped: {msg}", DANGER)
            return   # this device stops here; every other device keeps running untouched

        current_fw = result.get("firmware") or "Unknown"
        self._cell_signal.emit(row, self.COL_CURRENT_FW, current_fw, "")

        if self._cancel_flag:
            self._cell_signal.emit(row, self.COL_STATUS, "⏹ Stopped before send", WARNING)
            return

        # Step 2: send FOTA command
        self._cell_signal.emit(row, self.COL_STATUS, "📡 Sending FOTA command …", "")
        ok, msg = self._send_fota_command(ip, port, payload, log_cb)
        if not ok:
            self._cell_signal.emit(row, self.COL_STATUS, f"❌ Send failed — stopped: {msg}", DANGER)
            return

        # Step 3: wait 30s (interruptible)
        for remaining in range(30, 0, -1):
            if self._cancel_flag:
                break
            self._cell_signal.emit(row, self.COL_STATUS, f"⏳ Waiting {remaining}s …", "")
            time.sleep(1)
        if self._cancel_flag:
            self._cell_signal.emit(row, self.COL_STATUS, "⏹ Stopped during wait", WARNING)
            return

        # Step 4: reread firmware version
        self._cell_signal.emit(row, self.COL_STATUS, "🔄 Verifying new FW …", "")
        ok, msg, result2 = hdlc_read_firmware_blocking(ip, port, log_cb)
        if not ok:
            self._cell_signal.emit(row, self.COL_LATEST_FW, "Unknown", WARNING)
            self._cell_signal.emit(row, self.COL_STATUS, f"⚠ Sent, but verify failed: {msg}", WARNING)
            return

        latest_fw = result2.get("firmware") or "Unknown"
        self._cell_signal.emit(row, self.COL_LATEST_FW, latest_fw, "")
        if latest_fw != current_fw:
            self._cell_signal.emit(row, self.COL_STATUS, "✅ Updated", SUCCESS)
        else:
            self._cell_signal.emit(row, self.COL_STATUS, "⚠ Unchanged", WARNING)

    # ── Retry wrapper — mirrors the main page's _run_with_retries ──
    def _run_with_retries(self, step_name, attempt_fn, log_cb, max_attempts=3):
        result = None
        for attempt in range(1, max_attempts + 1):
            if self._cancel_flag:
                return result if result is not None else (False, "Cancelled", None)
            if attempt > 1:
                log_cb(f"🔁 '{step_name}' failed — retry {attempt}/{max_attempts} …")
                time.sleep(2)
            try:
                result = attempt_fn()
            except Exception as e:
                result = (False, f"{type(e).__name__}: {e}", None)
            if result[0]:
                return result
            log_cb(f"⚠ '{step_name}' attempt {attempt}/{max_attempts} failed — {result[1]}")
        return result

    def _process_one_device(self, row, ip, port, payload):
        if self._cancel_flag:
            self._cell_signal.emit(row, self.COL_STATUS, "⏹ Skipped (stopped)", WARNING)
            return

        def log_cb(msg):
            self._log_signal.emit(f"[{ip}] {msg}\n")

        # Step 1: read current firmware — retried 3x, same as the single-device page
        self._cell_signal.emit(row, self.COL_STATUS, "⏳ Reading current FW …", "")
        ok, msg, result = self._run_with_retries(
            "Read current firmware",
            lambda: hdlc_read_firmware_blocking(ip, port, log_cb),
            log_cb)
        if not ok:
            self._cell_signal.emit(row, self.COL_CURRENT_FW, "Unreachable", DANGER)
            self._cell_signal.emit(row, self.COL_STATUS, f"❌ Unreachable after 3 attempts: {msg}", DANGER)
            return   # this device stops here; every other device keeps running untouched

        current_fw = result.get("firmware") or "Unknown"
        self._cell_signal.emit(row, self.COL_CURRENT_FW, current_fw, "")

        if self._cancel_flag:
            self._cell_signal.emit(row, self.COL_STATUS, "⏹ Stopped before send", WARNING)
            return

        # Step 2: send FOTA command (already retries internally — unchanged)
        self._cell_signal.emit(row, self.COL_STATUS, "📡 Sending FOTA command …", "")
        ok, msg = self._send_fota_command(ip, port, payload, log_cb)
        if not ok:
            self._cell_signal.emit(row, self.COL_STATUS, f"❌ Send failed — stopped: {msg}", DANGER)
            return

        # Step 3: wait 30s (interruptible)
        for remaining in range(30, 0, -1):
            if self._cancel_flag:
                break
            self._cell_signal.emit(row, self.COL_STATUS, f"⏳ Waiting {remaining}s …", "")
            time.sleep(1)
        if self._cancel_flag:
            self._cell_signal.emit(row, self.COL_STATUS, "⏹ Stopped during wait", WARNING)
            return

        # Step 4: reread firmware version — also retried 3x
        self._cell_signal.emit(row, self.COL_STATUS, "🔄 Verifying new FW …", "")
        ok, msg, result2 = self._run_with_retries(
            "Reread firmware",
            lambda: hdlc_read_firmware_blocking(ip, port, log_cb),
            log_cb)
        if not ok:
            self._cell_signal.emit(row, self.COL_LATEST_FW, "Unknown", WARNING)
            self._cell_signal.emit(row, self.COL_STATUS, f"⚠ Sent, but verify failed after 3 attempts: {msg}", WARNING)
            return

        latest_fw = result2.get("firmware") or "Unknown"
        self._cell_signal.emit(row, self.COL_LATEST_FW, latest_fw, "")
        if latest_fw != current_fw:
            self._cell_signal.emit(row, self.COL_STATUS, "✅ Updated", SUCCESS)
        else:
            self._cell_signal.emit(row, self.COL_STATUS, "⚠ Unchanged", WARNING)

    def _device_finished(self):
        with self._active_lock:
            self._active_count -= 1
            done = self._active_count <= 0
        if done:
            self._append_log("=" * 60 + "\n")
            self._append_log(getattr(self, "_finish_message", "Bulk FOTA run complete — all devices finished.") + "\n")
            self._finished_signal.emit()

    def _send_fota_command(self, ip, port, payload, log_cb, max_attempts=3):
        """Actually opens a fresh TCP connection to the NIC and writes the FOTA
        command bytes on the wire. Retries a couple of times on transient
        connection errors (common over GPRS) before giving up on this device."""
        last_error = None

        for attempt in range(1, max_attempts + 1):
            try:
                sa = resolve_ipv6_sockaddr(ip, port)
                log_cb(f"[attempt {attempt}/{max_attempts}] Connecting to [{sa[0]}]:{sa[1]} …")

                with make_tcp_socket(timeout=10) as sock:
                    sock.connect(sa)
                    log_cb("TCP connection established.")
                    log_cb(f"RAW → {payload!r}")
                    sock.sendall(payload)
                    log_cb(f"✔ {len(payload)} bytes written to socket — command delivered over TCP.")

                    try:
                        resp = sock.recv(512).decode(errors="replace").strip()
                        if resp:
                            log_cb(f"NIC response: {resp}")
                    except (socket.timeout, ConnectionResetError):
                        log_cb("No immediate response (normal — NIC is downloading firmware)")

                return True, "FOTA command sent ✔"

            except ConnectionResetError:
                log_cb("Connection closed by NIC (expected — rebooting to apply firmware)")
                return True, "FOTA command sent (NIC closed connection — expected)"

            except (socket.timeout, ConnectionRefusedError, OSError) as e:
                last_error = f"{type(e).__name__}: {e}"
                log_cb(f"⚠ {last_error}")
                if attempt < max_attempts:
                    log_cb(f"Retrying in 2s …")
                    time.sleep(2)

            except Exception as e:
                # Non-network error — no point retrying
                return False, f"{type(e).__name__}: {e}"

        return False, last_error or "Connection failed after retries"

    def _on_finished(self):
        self._running = False
        self.run_btn.setEnabled(True)
        self.read_fw_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.upload_btn.setEnabled(True)

# ─────────────────────────────────────────────────────────────
#  FOTA Step Progress Dialog — driven by REAL step results
# ─────────────────────────────────────────────────────────────
class FotaStepDialog(QtWidgets.QDialog):
    STATUS_ICON = {"pending": "⚪", "running": "⏳", "ok": "✅", "fail": "❌", "warn": "⚠️"}

    def __init__(self, parent, steps):
        super().__init__(parent)
        self.setWindowTitle("FOTA Update — Live Progress")
        self.resize(900, 650)
        self.setMinimumWidth(900)
        self.setModal(True)
        self.setStyleSheet(f"QDialog {{ background-color: {BG_MAIN}; }}")

        self._cancelled = False
        self._done      = False
        self._steps     = steps

        root = QtWidgets.QVBoxLayout(self)
        root.setContentsMargins(22, 22, 22, 20)
        root.setSpacing(10)

        title = Label("🚀  FOTA Update in Progress", font=FONT_H1)
        title.setAlignment(Qt.AlignCenter)
        root.addWidget(title)
        sub = Label("Please wait — all controls are locked until this finishes.",
                     font=FONT_SMALL, color=MUTED_TEXT)
        sub.setAlignment(Qt.AlignCenter)
        root.addWidget(sub)

        self.overall_bar = ProgressBar(height=14)
        root.addWidget(self.overall_bar)
        root.addSpacing(5)
        self.overall_pct = Label("0 / " + str(len(steps)) + " steps complete",
                                  font=FONT_SMALL, color=MUTED_TEXT)
        self.overall_pct.setAlignment(Qt.AlignCenter)
        root.addWidget(self.overall_pct)

        steps_card = Card(self)
        steps_l = QtWidgets.QVBoxLayout(steps_card)
        steps_l.setContentsMargins(14, 10, 14, 10)
        steps_l.setSpacing(9)

        self._row_widgets = []
        for i, label in enumerate(steps):

            row = QtWidgets.QHBoxLayout()
            row.setSpacing(12)

            # Status Icon
            icon = Label("⚪", font=lambda: qfont(FONT_FAMILY, 15))
            icon.setFixedWidth(32)
            icon.setAlignment(Qt.AlignCenter)
            row.addWidget(icon)

            # Step Name
            text = Label(str(i + 1) + ". " + label, font=FONT_BODY)
            text.setMinimumWidth(360)
            text.setMaximumWidth(360)
            text.setWordWrap(False)
            text.setAlignment(Qt.AlignVCenter | Qt.AlignLeft)
            row.addWidget(text)

            # Status / Detail
            detail = Label("", font=FONT_SMALL, color=MUTED_TEXT)
            detail.setWordWrap(False)
            detail.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            row.addWidget(detail, 1)

            steps_l.addLayout(row)

            self._row_widgets.append((icon, text, detail))        
        
        root.addWidget(steps_card)

        root.addWidget(Label("Live Log", font=FONT_LABEL))
        self.log_box = TextBox(mono=True, height=140, readonly=True)
        root.addWidget(self.log_box)

        btn_row = QtWidgets.QHBoxLayout()
        btn_row.addStretch(1)
        self.close_btn = Button("Cancel", fg_color=DANGER, hover_color=DANGER_HOVER, width=130)
        self.close_btn.clicked.connect(self._cancel)
        btn_row.addWidget(self.close_btn)
        btn_row.addStretch(1)
        root.addLayout(btn_row)

    def closeEvent(self, event):
        # mirror the Tk version: block the window-close button until done
        if not self._done:
            event.ignore()
        else:
            event.accept()

    def _cancel(self):
        self._cancelled = True
        self.close_btn.setEnabled(False)
        self.close_btn.setText("Cancelling …")

    def _log(self, text):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_box.append_line("[" + ts + "] " + text + "\n")

    def set_step_status(self, i, status, detail=None):
        icon, text, detail_lbl = self._row_widgets[i]
        icon.setText(self.STATUS_ICON.get(status, "⚪"))
        color_map = {"running": WARNING, "ok": SUCCESS, "fail": DANGER,
                     "pending": TEXT_DARK, "warn": WARNING}
        text.set_color(color_map.get(status, TEXT_DARK))
        if detail:
            detail_lbl.setText(detail)

        done_count = sum(1 for ic, _, _ in self._row_widgets
                          if ic.text() == self.STATUS_ICON["ok"])
        total = len(self._row_widgets)
        self.overall_bar.set_fraction(done_count / total if total else 0)
        self.overall_pct.setText(str(done_count) + " / " + str(total) + " steps complete")

        self._log(self._steps[i] + " — " + status.upper() + (("  (" + detail + ")") if detail else ""))

    def update_wait_countdown(self, i, remaining):
        _, _, detail_lbl = self._row_widgets[i]
        detail_lbl.setText(str(remaining) + "s remaining …")

    def mark_failed(self, i, reason):
        self.set_step_status(i, "fail", reason)
        self.overall_bar.set_color(DANGER)
        self.close_btn.setText("Close")
        self.close_btn.set_colors(DANGER, DANGER_HOVER)
        self.close_btn.setEnabled(True)
        try:
            self.close_btn.clicked.disconnect()
        except TypeError:
            pass
        self.close_btn.clicked.connect(self.accept)
        self._done = True

    def mark_cancelled(self):
        self.overall_bar.set_color(WARNING)
        self._log("⚠ Cancelled by user.")
        self.close_btn.setText("Close")
        self.close_btn.set_colors(WARNING, WARNING_HOVER)
        self.close_btn.setEnabled(True)
        try:
            self.close_btn.clicked.disconnect()
        except TypeError:
            pass
        self.close_btn.clicked.connect(self.accept)
        self._done = True

    def mark_all_complete(self, old_ver, new_ver):
        self.overall_bar.set_fraction(1.0)
        self.overall_bar.set_color(SUCCESS)
        self._log("✅ FOTA sequence complete.  " + (old_ver or "?") + "  →  " + (new_ver or "?"))
        self.close_btn.setText("Close")
        self.close_btn.set_colors(SUCCESS, SUCCESS_HOVER)
        self.close_btn.setEnabled(True)
        try:
            self.close_btn.clicked.disconnect()
        except TypeError:
            pass
        self.close_btn.clicked.connect(self.accept)
        self._done = True


# ─────────────────────────────────────────────────────────────
#  Main App
# ─────────────────────────────────────────────────────────────
class OrionFOTAApp(QtWidgets.QMainWindow):

    FOTA_STEPS = [
        "Validate firmware URL",
        "Validate NIC IPv6 address",
        "Verify firmware file is accessible",
        "Ping test the NIC",
        "Read current firmware version",
        "Send FOTA command",
        "Close socket & wait for NIC to update (30s)",
        "Reconnect & verify firmware version",
    ]

    def __init__(self):
        super().__init__()
        self.setWindowTitle("⚡ ORION EC200 / EC800 Firmware Updater")
        self.resize(1240, 920)
        self.setMinimumSize(1040, 740)
        self.setStyleSheet(f"QMainWindow {{ background-color: {BG_MAIN}; }}")

        self._dispatch = _Dispatcher()

        self._fota_running = False
        self._fw_btn_busy  = False
        self._settings     = load_settings()
        self._freezable    = []   # every control that must lock during FOTA

        self.build_ui()
        self._restore_settings()

    # ── thread → UI-thread dispatch (mirrors Tk's self.after(0, fn, *a)) ──
    def after(self, _ms, fn, *args, **kwargs):
        self._dispatch.fire.emit(fn, args, kwargs)

    def closeEvent(self, event):
        if self._fota_running:
            if not mb.askyesno("FOTA in progress",
                                "A FOTA update is currently running.\n\n"
                                "Closing now may interrupt the update. Close anyway?"):
                event.ignore()
                return
        self._save_current_settings()
        event.accept()

    # ── Settings ──────────────────────────────────────────────
    def _save_current_settings(self):
        data = {
            "ip":       self.ip_entry.text().strip(),
            "port":     self.nic_port_entry.text().strip(),
            "protocol": "TCP" if self.rb_tcp.isChecked() else "UDP",
            "fota": {label: self._entry_value(e) for label, e in self.entries.items()
                     if e is not None},
            "fp_url":   self.fp_url_entry.text().strip(),
            "fp_func":  self.fp_func_entry.text().strip(),
            "fp_fname": self.fp_fname_entry.text().strip(),
        }
        save_settings(data)

    @staticmethod
    def _entry_value(e):
        if isinstance(e, QtWidgets.QComboBox):
            return e.currentText()
        return e.text()

    @staticmethod
    def _set_entry_value(e, val):
        if isinstance(e, QtWidgets.QComboBox):
            idx = e.findText(val)
            if idx >= 0:
                e.setCurrentIndex(idx)
        else:
            e.setText(val)

    def _restore_settings(self):
        s = self._settings
        if not s:
            return
        if "ip" in s:
            self.ip_entry.setText(s["ip"])
        if "port" in s:
            self.nic_port_entry.setText(s["port"])
        if s.get("protocol") == "UDP":
            self.rb_udp.setChecked(True)
        else:
            self.rb_tcp.setChecked(True)
        fota = s.get("fota", {})
        for label, val in fota.items():
            if label in self.entries and self.entries[label] is not None:
                self._set_entry_value(self.entries[label], val)
        if "fp_url" in s:
            self.fp_url_entry.setText(s["fp_url"])
        if "fp_func" in s:
            self.fp_func_entry.setText(s["fp_func"])
        if "fp_fname" in s:
            self.fp_fname_entry.setText(s["fp_fname"])
        self.update_preview()

    # ── UI ────────────────────────────────────────────────────
    def build_ui(self):
        central = QtWidgets.QWidget()
        self.setCentralWidget(central)
        outer = QtWidgets.QVBoxLayout(central)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        # ── Header banner ──
        header = QtWidgets.QFrame()
        header.setFixedHeight(78)
        header.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {BRAND}, stop:1 {BRAND_HOVER});
            }}
        """)
        h_l = QtWidgets.QHBoxLayout(header)
        h_l.setContentsMargins(24, 10, 24, 10)

        title_box = QtWidgets.QVBoxLayout()
        t1 = Label("⚡  ORION Firmware Updater", font=lambda: qfont(FONT_FAMILY, 18, QtGui.QFont.Bold), color="white")
        title_box.addWidget(t1)
        t2 = Label("📡 EC200 / EC800  ·  DLMS/COSEM NIC over IPv6", font=FONT_SMALL, color="#eef3e2")
        title_box.addWidget(t2)
        h_l.addLayout(title_box)
        h_l.addStretch(1)

        self.status_pill = StatusPill("●  Idle", bg="#8a9c5c")
        h_l.addWidget(self.status_pill, alignment=Qt.AlignVCenter)
        outer.addWidget(header)

        # ── Scrollable body ──
        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        body = QtWidgets.QWidget()
        body.setStyleSheet(f"background-color: {BG_MAIN};")
        body_l = QtWidgets.QVBoxLayout(body)
        body_l.setContentsMargins(16, 16, 16, 16)
        body_l.setSpacing(12)
        scroll.setWidget(body)
        outer.addWidget(scroll, 1)

        # ── Firmware version row ──
        fw_card = Card()
        fw_l = QtWidgets.QHBoxLayout(fw_card)
        fw_l.setContentsMargins(16, 14, 16, 14)
        fw_l.addWidget(Label("🔧  Current Firmware Version", font=FONT_LABEL))
        self.version_var = StringVar("Not Fetched")
        self.version_label = Label(textvariable=self.version_var,
                                    font=lambda: qfont(FONT_FAMILY, 16, QtGui.QFont.Bold),
                                    color=BRAND_HOVER)
        fw_l.addWidget(self.version_label)
        fw_l.addStretch(1)

        self.get_version_btn = Button("Get Version", fg_color=SLATE, hover_color=SLATE_HOVER, width=110)
        self.get_version_btn.clicked.connect(self.get_version)
        fw_l.addWidget(self.get_version_btn)
        self._freezable.append(self.get_version_btn)

        self.nic_details_btn = Button("ℹ️  Get NIC Details", fg_color=INFO, hover_color=INFO_HOVER, width=160)
        self.nic_details_btn.clicked.connect(self.get_nic_details)
        fw_l.addWidget(self.nic_details_btn)
        self._freezable.append(self.nic_details_btn)

        self.nic_reset_btn = Button("🔄  NIC Reset", fg_color=WARNING, hover_color=WARNING_HOVER, width=130)
        self.nic_reset_btn.clicked.connect(self.send_nic_reset)
        fw_l.addWidget(self.nic_reset_btn)
        self._freezable.append(self.nic_reset_btn)

        self.nic_fw_btn = Button("📟  Get NIC Firmware", fg_color=INFO, hover_color=INFO_HOVER, width=180)
        self.nic_fw_btn.clicked.connect(self.get_nic_firmware)
        fw_l.addWidget(self.nic_fw_btn)
        self._freezable.append(self.nic_fw_btn)

        body_l.addWidget(fw_card)

        # ── Device config ──
        dev_card = Card()
        dev_l = QtWidgets.QGridLayout(dev_card)
        dev_l.setContentsMargins(16, 14, 16, 14)
        dev_l.setVerticalSpacing(10)
        dev_l.setHorizontalSpacing(8)

        dev_l.addWidget(Label("📶  Protocol", font=FONT_LABEL), 0, 0)
        proto_row = QtWidgets.QHBoxLayout()
        self.rb_tcp = QtWidgets.QRadioButton("TCP")
        self.rb_udp = QtWidgets.QRadioButton("UDP")
        self.rb_tcp.setChecked(True)
        for rb in (self.rb_tcp, self.rb_udp):
            rb.setFont(FONT_BODY())
            rb.setStyleSheet(f"QRadioButton::indicator:checked {{ background-color: {BRAND}; }}")
        proto_row.addWidget(self.rb_tcp)
        proto_row.addWidget(self.rb_udp)
        proto_row.addStretch(1)
        dev_l.addLayout(proto_row, 0, 1, 1, 3)
        self._freezable += [self.rb_tcp, self.rb_udp]

        dev_l.addWidget(Label("🌐  NIC IPv6 Address", font=FONT_LABEL), 1, 0)
        self.ip_entry = LineEdit("2401:4900:984a:3399::2")
        dev_l.addWidget(self.ip_entry, 1, 1)
        self._freezable.append(self.ip_entry)

        dev_l.addWidget(Label("Port", font=FONT_LABEL), 1, 2)
        self.nic_port_entry = LineEdit("4059", width=90)
        dev_l.addWidget(self.nic_port_entry, 1, 3)
        self._freezable.append(self.nic_port_entry)

        self.ping_btn_top = Button("🔔  Ping", fg_color=SLATE, hover_color=SLATE_HOVER, width=100)
        self.ping_btn_top.clicked.connect(self.open_ping_popup)
        dev_l.addWidget(self.ping_btn_top, 1, 4)
        self._freezable.append(self.ping_btn_top)
        dev_l.setColumnStretch(1, 1)
        body_l.addWidget(dev_card)

        # ── FOTA config (2-column layout) ───────────────────────────────
        fota_card = Card()
        fota_l = QtWidgets.QGridLayout(fota_card)
        fota_l.setContentsMargins(16, 14, 16, 14)
        fota_l.setVerticalSpacing(8)
        fota_l.setHorizontalSpacing(6)

        fota_l.addWidget(Label("🛰️  FOTA Command Parameters", font=FONT_H2), 0, 0, 1, 4)

        self.entries = {}

        # ---------------- Left Side ----------------
        left_col = [
            ("User", "0"),
            ("Password", "0"),
            ("Server IP", "0"),
            ("File Port", "0")
        ]

        for row, (label, default) in enumerate(left_col, start=1):

            lbl = Label(
                label,
                font=lambda: qfont(FONT_FAMILY, 10, QtGui.QFont.Medium)
            )
            fota_l.addWidget(lbl, row, 0)

            entry = LineEdit(default)
            entry.setSizePolicy(
                QtWidgets.QSizePolicy.Expanding,
                QtWidgets.QSizePolicy.Fixed
            )
            entry.textChanged.connect(lambda _t: self.update_preview())

            fota_l.addWidget(entry, row, 1)

            self.entries[label] = entry
            self._freezable.append(entry)

        # FOTA Mode
        lbl = Label(
            "FOTA Mode",
            font=lambda: qfont(FONT_FAMILY, 10, QtGui.QFont.Medium)
        )
        fota_l.addWidget(lbl, 5, 0)

        mode_combo = ComboBox(
            values=[
                "0 - Full FOTA",
                "1 - Delta FOTA"
            ]
        )
        mode_combo.setSizePolicy(
            QtWidgets.QSizePolicy.Expanding,
            QtWidgets.QSizePolicy.Fixed
        )
        mode_combo.currentTextChanged.connect(lambda _t: self.update_preview())

        fota_l.addWidget(mode_combo, 5, 1)

        self.entries["FOTA Mode"] = mode_combo
        self._freezable.append(mode_combo)

        # ---------------- Right Side ----------------
        _default_url = "https://ties-nancy-mary-dos.trycloudflare.com"
        _default_func = "getFile"
        _default_fname = "Full_FOTA_13X.pac"

        # File URL
        lbl = Label(
            "File URL",
            font=lambda: qfont(FONT_FAMILY, 10, QtGui.QFont.Medium)
        )
        fota_l.addWidget(lbl, 1, 2)

        self.fp_url_entry = LineEdit(_default_url)
        self.fp_url_entry.setPlaceholderText("https://example.com")
        self.fp_url_entry.setSizePolicy(
            QtWidgets.QSizePolicy.Expanding,
            QtWidgets.QSizePolicy.Fixed
        )
        self.fp_url_entry.textChanged.connect(lambda _t: self.update_preview())
        fota_l.addWidget(self.fp_url_entry, 1, 3)
        self._freezable.append(self.fp_url_entry)

        # Function
        lbl = Label(
            "Function",
            font=lambda: qfont(FONT_FAMILY, 10, QtGui.QFont.Medium)
        )
        fota_l.addWidget(lbl, 2, 2)

        self.fp_func_entry = LineEdit(_default_func)
        self.fp_func_entry.setPlaceholderText("getFile")
        self.fp_func_entry.setSizePolicy(
            QtWidgets.QSizePolicy.Expanding,
            QtWidgets.QSizePolicy.Fixed
        )
        self.fp_func_entry.textChanged.connect(lambda _t: self.update_preview())
        fota_l.addWidget(self.fp_func_entry, 2, 3)
        self._freezable.append(self.fp_func_entry)

        # File Name
        lbl = Label(
            "File Name",
            font=lambda: qfont(FONT_FAMILY, 10, QtGui.QFont.Medium)
        )
        fota_l.addWidget(lbl, 3, 2)

        self.fp_fname_entry = LineEdit(_default_fname)
        self.fp_fname_entry.setPlaceholderText("firmware.pac")
        self.fp_fname_entry.setSizePolicy(
            QtWidgets.QSizePolicy.Expanding,
            QtWidgets.QSizePolicy.Fixed
        )
        self.fp_fname_entry.textChanged.connect(lambda _t: self.update_preview())
        fota_l.addWidget(self.fp_fname_entry, 3, 3)
        self._freezable.append(self.fp_fname_entry)

        self.entries["File Path"] = None

        # ---------------- Layout Stretch ----------------

        # Small width for labels
        fota_l.setColumnMinimumWidth(0, 80)
        fota_l.setColumnMinimumWidth(2, 90)

        # Large width for input fields
        fota_l.setColumnStretch(0, 0)
        fota_l.setColumnStretch(1, 4)
        fota_l.setColumnStretch(2, 0)
        fota_l.setColumnStretch(3, 6)

        body_l.addWidget(fota_card)

        # ── Command preview ──
        preview_card = Card()
        prev_l = QtWidgets.QVBoxLayout(preview_card)
        prev_l.setContentsMargins(16, 12, 16, 12)
        prev_l.addWidget(Label("👁️  Command Preview", font=FONT_LABEL))
        self.command_var = StringVar()
        self.command_preview = LineEdit(mono=True)
        self.command_preview.setReadOnly(True)
        self.command_var.changed.connect(self.command_preview.setText)
        prev_l.addWidget(self.command_preview)
        body_l.addWidget(preview_card)
        self.update_preview()

        # ── Action buttons ──
        action_card = Card()
        act_l = QtWidgets.QHBoxLayout(action_card)
        act_l.setContentsMargins(16, 14, 16, 14)

        self.test_btn = Button("🔗  Test Connection", fg_color=INFO, hover_color=INFO_HOVER, width=170)
        self.test_btn.clicked.connect(self.test_connection)
        act_l.addWidget(self.test_btn)
        self._freezable.append(self.test_btn)

        self.fota_btn = Button("🚀  Send FOTA", fg_color=BRAND, hover_color=BRAND_HOVER,
                                width=160, height=40, font=lambda: qfont(FONT_FAMILY, 12, QtGui.QFont.Bold))
        self.fota_btn.clicked.connect(self.send_fota)
        act_l.addWidget(self.fota_btn)
        self._freezable.append(self.fota_btn)

        self.reset_btn = Button("♻️  Reset", fg_color=SLATE, hover_color=SLATE_HOVER, width=110)
        self.reset_btn.clicked.connect(self.reset_all)
        act_l.addWidget(self.reset_btn)
        self._freezable.append(self.reset_btn)

        self.ping_btn_bottom = Button("🔔  Ping", fg_color=SLATE, hover_color=SLATE_HOVER, width=100)
        self.ping_btn_bottom.clicked.connect(self.open_ping_popup)
        act_l.addWidget(self.ping_btn_bottom)
        self._freezable.append(self.ping_btn_bottom)

        self.stress_btn = Button("🔥  Stress Test", fg_color=PURPLE, hover_color=PURPLE_HOVER, width=140)
        self.stress_btn.clicked.connect(self.open_stress_test_popup)
        act_l.addWidget(self.stress_btn)
        self._freezable.append(self.stress_btn)

        self.bulk_fota_btn = Button("🛰️  Bulk FOTA", fg_color=INFO, hover_color=INFO_HOVER, width=140)
        self.bulk_fota_btn.clicked.connect(self.open_bulk_fota_popup)
        act_l.addWidget(self.bulk_fota_btn)
        self._freezable.append(self.bulk_fota_btn)

        act_l.addStretch(1)
        body_l.addWidget(action_card)

        # ── Log area ──
        log_card = Card()
        log_l = QtWidgets.QVBoxLayout(log_card)
        log_l.setContentsMargins(16, 12, 16, 14)
        log_header = QtWidgets.QHBoxLayout()
        log_header.addWidget(Label("📝  Logs", font=FONT_LABEL))
        log_header.addStretch(1)
        self.clear_log_btn = Button("Clear Logs", fg_color=SLATE, hover_color=SLATE_HOVER, width=100, height=28)
        self.clear_log_btn.clicked.connect(lambda: self.log_box.clear_all())
        log_header.addWidget(self.clear_log_btn)
        self._freezable.append(self.clear_log_btn)
        log_l.addLayout(log_header)

        self.log_box = TextBox(mono=True, height=240, readonly=True)
        log_l.addWidget(self.log_box)
        body_l.addWidget(log_card, 1)

        # ── Status bar ──
        self.status_var = StringVar("Disconnected")
        status_bar = QtWidgets.QStatusBar()
        status_bar.setStyleSheet(f"QStatusBar {{ background-color: #e9ede3; color: #444; }}")
        self.status_bar_label = QtWidgets.QLabel("Disconnected")
        self.status_bar_label.setFont(FONT_SMALL())
        self.status_var.changed.connect(self.status_bar_label.setText)
        status_bar.addWidget(self.status_bar_label, 1)
        self.setStatusBar(status_bar)

    # ── helpers ──────────────────────────────────────────────
    def _get_file_path(self) -> str:
        url   = self.fp_url_entry.text().strip().rstrip("/")
        func  = self.fp_func_entry.text().strip().strip("/")
        fname = self.fp_fname_entry.text().strip()
        if func:
            return url + "/" + func + "/" + fname
        return url + "/" + fname

    def update_preview(self):
        mode_raw = self.entries["FOTA Mode"].currentText()
        mode_val = mode_raw.split(" ")[0]
        fname = self.fp_fname_entry.text().strip()

        if mode_val == "1":
            if fname.endswith(".pac") and not fname.endswith(".pack"):
                fname = fname[:-4] + ".pack"
        else:
            if fname.endswith(".pack"):
                fname = fname[:-5] + ".pac"

        url  = self.fp_url_entry.text().strip().rstrip("/")
        func = self.fp_func_entry.text().strip().strip("/")
        file_path = url + "/" + func + "/" + fname if func else url + "/" + fname

        cmd = (
            "*GPRS-FOTA,"
            + self.entries["User"].text() + ","
            + self.entries["Password"].text() + ","
            + self.entries["Server IP"].text() + ","
            + mode_val + ","
            + file_path + ","
            + self.entries["File Port"].text() + ",#"
        )
        self.command_var.set(cmd)

    def log(self, message):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_box.append_line("[" + ts + "] " + message + "\n")

    def _set_status_pill(self, text, color="#8a9c5c"):
        self.status_pill.set_status(text, color)

    def _get_ip_port(self):
        ip   = self.ip_entry.text().strip()
        port = int(self.nic_port_entry.text().strip() or 4059)
        return ip, port

    def open_ping_popup(self):
        dlg = PingDialog(self, default_ip=self.ip_entry.text().strip())
        dlg.exec()

    def open_stress_test_popup(self):
        ip, port = self._get_ip_port()
        dlg = StressTestDialog(self, ip=ip, port=port)
        dlg.exec()

    def open_bulk_fota_popup(self):
        ip, port = self._get_ip_port()
        dlg = BulkFotaDialog(self, default_ip_port=(ip, port))
        dlg.exec()

    # ── Freeze / unfreeze every control during FOTA ─────────────
    def _freeze_ui(self):
        for w in self._freezable:
            try:
                w.setEnabled(False)
            except Exception:
                pass

    def _unfreeze_ui(self):
        for w in self._freezable:
            try:
                w.setEnabled(True)
            except Exception:
                pass

    # ── Get NIC Firmware ────────────────────────────────
    def get_nic_firmware(self):
        if self._fw_btn_busy or self._fota_running:
            return
        self._fw_btn_busy = True
        self.nic_fw_btn.setEnabled(False)
        self.nic_fw_btn.setText("⏳  Reading …")
        ip, port = self._get_ip_port()
        self.log("=" * 60)
        self.log(f"📟  Getting NIC firmware from [{ip}]:{port} …")
        self.log("    Sending HDLC sequence: SNRM → AARQ → RLRQ → NIC Firmware → DISC")

        def _log_cb(msg):
            self.after(0, self.log, msg)

        def _done_cb(result, error_msg=None):
            self.after(0, self._on_nic_firmware_done, result, error_msg)

        run_nic_firmware_sequence(ip, port, _log_cb, _done_cb)

    def _on_nic_firmware_done(self, result, error_msg=None):
        self._fw_btn_busy = False
        self.nic_fw_btn.setEnabled(True)
        self.nic_fw_btn.setText("📟  Get NIC Firmware")

        if result is None:
            reason = error_msg or "Unknown error — check logs above."
            self.log(f"❌  NIC Firmware retrieval failed — {reason}")
            self.status_var.set("✘  NIC firmware read failed")
            mb.showerror("Could Not Retrieve Firmware Version",
                         "Could not retrieve the NIC firmware version.\n\n" + reason)
            return

        fw  = result.get("firmware") or "—"
        imei = result.get("imei")    or "—"
        self.log(f"✅  NIC Firmware: {fw}   |   IMEI: {imei}")
        self.status_var.set(f"✅  NIC FW: {fw}  |  IMEI: {imei}")
        NicFirmwareDialog(self, result, error_msg).exec()

    # ── Generic raw AT-style command sender (used by Reset / Details) ──
    def _send_raw_gprs_command(self, cmd: str, label: str):
        ip, port = self._get_ip_port()
        payload = (cmd + "\r\n").encode("ascii")
        try:
            sa = resolve_ipv6_sockaddr(ip, port)
            self.after(0, self.log, f"{label}  Connecting to [{sa[0]}]:{sa[1]} …")
            with make_tcp_socket(timeout=10) as sock:
                sock.connect(sa)
                self.after(0, self.log, f"{label}  RAW → {payload!r}")
                sock.sendall(payload)
                self.after(0, self.log, f"{label}  Sent {len(payload)} bytes ✔")
                try:
                    resp = sock.recv(1024).decode(errors="replace").strip()
                except (socket.timeout, ConnectionResetError):
                    resp = ""
            if resp:
                self.after(0, self.log, f"{label}  NIC response: {resp}")
                return True, resp
            return True, "(no response received — command sent successfully)"
        except Exception as e:
            reason = type(e).__name__ + ": " + str(e)
            self.after(0, self.log, f"{label}  ❌ {reason}")
            return False, reason

    # ── NIC Reset (*GPRS-RESET) ─────────────────────────────────
    def send_nic_reset(self):
        if self._fota_running:
            return
        if not mb.askyesno("Confirm NIC Reset",
                            "This will send *GPRS-RESET# to the NIC and reset it.\n\n"
                            "Continue?"):
            return
        self.nic_reset_btn.setEnabled(False)
        self.nic_reset_btn.setText("⏳  Resetting …")
        self.log("=" * 60)
        self.log("🔄  Sending NIC Reset command …")

        def _worker():
            ok, resp = self._send_raw_gprs_command("*GPRS-RESET#", "RESET")
            self.after(0, self._on_nic_reset_done, ok, resp)

        threading.Thread(target=_worker, daemon=True).start()

    def _on_nic_reset_done(self, ok, resp):
        self.nic_reset_btn.setEnabled(True)
        self.nic_reset_btn.setText("🔄  NIC Reset")
        if ok:
            mb.showinfo("NIC Reset", "Command sent: *GPRS-RESET#\n\nResponse:\n" + resp)
        else:
            mb.showerror("NIC Reset Failed", "Could not send *GPRS-RESET#.\n\n" + resp)

    # ── Get NIC Details (*GPRS-INFO) ────────────────────────────
    def get_nic_details(self):
        if self._fota_running:
            return
        self.nic_details_btn.setEnabled(False)
        self.nic_details_btn.setText("⏳  Reading …")
        self.log("=" * 60)
        self.log("ℹ️  Requesting NIC details …")

        def _worker():
            ok, resp = self._send_raw_gprs_command("*GPRS-INFO#", "INFO")
            self.after(0, self._on_nic_details_done, ok, resp)

        threading.Thread(target=_worker, daemon=True).start()

    def _on_nic_details_done(self, ok, resp):
        self.nic_details_btn.setEnabled(True)
        self.nic_details_btn.setText("ℹ️  Get NIC Details")
        if ok:
            mb.showinfo("NIC Details", "Command sent: *GPRS-INFO#\n\nResponse:\n" + resp)
            fw_slice = resp[21:37].strip()   # 22nd–38th characters (1-indexed)
            if fw_slice:
                self.version_var.set(fw_slice)
        else:
            mb.showerror("Get NIC Details Failed", "Could not send *GPRS-INFO#.\n\n" + resp)

    # ── Reset ─────────────────────────────────────────────────
    def reset_all(self):
        if self._fota_running:
            self.log("⚠  Cannot reset while FOTA is in progress.")
            return
        self.log_box.clear_all()
        self.version_var.set("Not Fetched")
        self.status_var.set("Disconnected")
        defaults = {"User": "0", "Password": "0", "Server IP": "0", "File Port": "0"}
        for label, val in defaults.items():
            self.entries[label].setText(val)
        idx = self.entries["FOTA Mode"].findText("0 - Full FOTA")
        if idx >= 0:
            self.entries["FOTA Mode"].setCurrentIndex(idx)
        self.fp_url_entry.setText("https://ties-nancy-mary-dos.trycloudflare.com")
        self.fp_func_entry.setText("getFile")
        self.fp_fname_entry.setText("Full_FOTA_13X.pac")
        self.update_preview()
        self.fota_btn.set_colors(BRAND, BRAND_HOVER)
        self.fota_btn.setText("🚀  Send FOTA")
        self.fota_btn.setEnabled(True)
        self._set_status_pill("●  Idle", "#8a9c5c")
        self.log("♻️  Reset complete.")
        self._save_current_settings()

    # ── Get Version ──────────────────────────────────────────
    def get_version(self):
        if self._fota_running:
            return
        self.log("=" * 60)
        self.log("🚀  Get Firmware Version requested …")
        self.version_var.set("Checking …")
        threading.Thread(target=self._get_version_worker, daemon=True).start()

    def _get_version_worker(self):
        ip, port = self._get_ip_port()
        sock = None
        try:
            self.after(0, self.log, "🔎  Checking NIC is reachable …")
            ok, msg, sa = self._check_nic_reachable(ip, port)
            if not ok:
                self.after(0, self.log, "❌  NIC not reachable — " + msg)
                self.after(0, self.version_var.set, "Not Reachable")
                return
            self.after(0, self.log, "✅  " + msg)

            self.after(0, self.log, "Connecting to " + sa[0] + " port " + str(sa[1]) + " …")
            sock = make_tcp_socket(timeout=5)
            sock.connect(sa)
            self.after(0, self.log, "Sending *GPRS-APPVER …")
            sock.sendall(b"*GPRS-APPVER\r\n")
            resp = sock.recv(256).decode(errors="replace").strip()
            if resp:
                self.after(0, self.version_var.set, resp)
                self.after(0, self.log, "✅  Version: " + resp)
            else:
                self.after(0, self.version_var.set, "No Response")
                self.after(0, self.log, "⚠️  NIC closed connection with no response.")
        except socket.timeout:
            self.after(0, self.log, "❌  Get Version timed out waiting for NIC response.")
            self.after(0, self.version_var.set, "Timeout")
        except Exception as e:
            self.after(0, self.log, "❌  Get Version failed — " + type(e).__name__ + ": " + str(e))
            self.after(0, self.version_var.set, "Error")
        finally:
            if sock is not None:
                try:
                    sock.close()
                    self.after(0, self.log, "🔌  Socket closed.")
                except Exception:
                    pass

    # ── Test Connection ──────────────────────────────────────
    def test_connection(self):
        if self._fota_running:
            return
        ip, port = self._get_ip_port()
        proto = "TCP" if self.rb_tcp.isChecked() else "UDP"
        self.log("Testing " + proto + " → [" + ip + "]:" + str(port) + " …")
        threading.Thread(target=self._test_worker, daemon=True).start()

    def _test_worker(self):
        ip, port = self._get_ip_port()
        proto = "TCP" if self.rb_tcp.isChecked() else "UDP"
        try:
            sa = resolve_ipv6_sockaddr(ip, port)
            self.after(0, self.log, "Resolved: " + sa[0] + " port " + str(sa[1]))
            if proto == "TCP":
                with make_tcp_socket(timeout=5) as sock:
                    sock.connect(sa)
                    try:
                        sock.sendall(b"\r\n")
                        banner = sock.recv(128).decode(errors="replace").strip()
                        if banner:
                            self.after(0, self.log, "Banner: " + banner)
                    except socket.timeout:
                        pass
                self.after(0, self.status_var.set, "✔ TCP Connected → [" + ip + "]:" + str(port))
                self.after(0, self.log, "TCP connection successful ✔")
            else:
                with make_udp_socket(timeout=3) as sock:
                    sock.connect(sa)
                self.after(0, self.status_var.set, "✔ UDP ready → [" + ip + "]:" + str(port))
                self.after(0, self.log, "UDP socket OK ✔")
        except socket.gaierror as e:
            self.after(0, self.log, "Address resolution failed: " + str(e))
            self.after(0, self.status_var.set, "✘ Address error")
        except ConnectionRefusedError:
            self.after(0, self.log, "Connection refused — NIC not listening on port " + str(port))
            self.after(0, self.status_var.set, "✘ Connection refused")
        except socket.timeout:
            self.after(0, self.log, "Timed out — check NIC reachable from this machine")
            self.after(0, self.status_var.set, "✘ Timeout")
        except OSError as e:
            self.after(0, self.log, "Socket OS error: " + str(e))
            self.after(0, self.status_var.set, "✘ Socket error")

    # ── Validation helpers ──────────────────────────────────
    def _validate_url_format(self, file_path: str):
        try:
            parsed = urlparse(file_path)
        except Exception as e:
            return False, "URL could not be parsed: " + str(e)
        if parsed.scheme not in ("http", "https"):
            return False, "URL must start with http:// or https://"
        if not parsed.netloc:
            return False, "URL is missing a host (e.g. https://example.com/...)"
        if not parsed.path or parsed.path == "/":
            return False, "URL has no file path / filename."
        return True, "URL is well-formed."

    def _check_remote_file_exists(self, file_path: str, timeout: int = 8):
        req = urllib.request.Request(file_path, method="HEAD")
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                if 200 <= resp.status < 400:
                    return True, "File reachable (HTTP " + str(resp.status) + ")"
                return False, "Server returned HTTP " + str(resp.status)
        except urllib.error.HTTPError as e:
            if e.code == 405:
                try:
                    req2 = urllib.request.Request(
                        file_path, method="GET",
                        headers={"Range": "bytes=0-0"})
                    with urllib.request.urlopen(req2, timeout=timeout) as resp2:
                        if 200 <= resp2.status < 400:
                            return True, "File reachable (HTTP " + str(resp2.status) + ")"
                        return False, "Server returned HTTP " + str(resp2.status)
                except urllib.error.HTTPError as e2:
                    return False, "File not found on server (HTTP " + str(e2.code) + ")"
                except Exception as e2:
                    return False, "Could not reach file URL: " + type(e2).__name__ + ": " + str(e2)
            if e.code == 404:
                return False, "File not found on server (HTTP 404)"
            return False, "Server returned HTTP " + str(e.code)
        except urllib.error.URLError as e:
            return False, "Could not reach URL host: " + str(e.reason)
        except Exception as e:
            return False, "URL check failed: " + type(e).__name__ + ": " + str(e)

    def _check_nic_reachable(self, ip: str, port: int, timeout: int = 5):
        if not ip:
            return False, "NIC IP address is empty.", None
        try:
            sa = resolve_ipv6_sockaddr(ip, port)
        except socket.gaierror as e:
            return False, "Could not resolve NIC IPv6 address: " + str(e), None
        except Exception as e:
            return False, "Address resolution failed: " + type(e).__name__ + ": " + str(e), None

        sock = None
        try:
            sock = make_tcp_socket(timeout=timeout)
            sock.connect(sa)
            return True, "NIC reachable at [" + sa[0] + "]:" + str(sa[1]), sa
        except ConnectionRefusedError:
            return False, "Connection refused -- NIC not listening on port " + str(port), sa
        except socket.timeout:
            return False, "Timed out connecting to NIC -- check it's powered on and reachable", sa
        except OSError as e:
            return False, "Socket error while connecting to NIC: " + str(e), sa
        finally:
            if sock is not None:
                try:
                    sock.close()
                except Exception:
                    pass

    def _ping_test_ip(self, ip: str, count: int = 2):
        is_win = platform.system().lower() == "windows"
        is_v6  = ":" in ip
        if is_v6:
            candidates = (["ping", "-6", "-n", str(count), ip] if is_win else
                          [["ping6", "-c", str(count), ip], ["ping", "-6", "-c", str(count), ip]])
            if is_win:
                candidates = [candidates]
        else:
            flag = "-n" if is_win else "-c"
            candidates = [["ping", flag, str(count), ip]]

        last_err = None
        for cmd in candidates:
            try:
                proc = subprocess.run(cmd, capture_output=True, text=True, timeout=6 * count + 6)
            except FileNotFoundError as e:
                last_err = str(e)
                continue
            except subprocess.TimeoutExpired:
                return False, "Ping command timed out."
            output = (proc.stdout or "") + (proc.stderr or "")
            ll = output.lower()

            has_reply   = any(k in ll for k in ("ttl=", "time=", "bytes from", "time<1ms"))
            has_unreach = any(k in ll for k in ("destination net unreachable",
                                                 "destination host unreachable",
                                                 "network is unreachable"))

            if has_reply:
                return True, "Host responded to ping."

            if has_unreach or "100% packet loss" in ll or "100% loss" in ll:
                return False, "Host did not respond to ping (unreachable)."

            if proc.returncode == 0:
                return True, "Ping succeeded."
            return False, "Host did not respond to ping (unreachable)."

        return False, "No suitable ping command found on this system." + \
            (" (" + last_err + ")" if last_err else "")

    def _validate_file_extension(self) -> bool:
        mode_raw = self.entries["FOTA Mode"].currentText()
        mode_val = mode_raw.split(" ")[0]
        fname    = self.fp_fname_entry.text().strip()

        if mode_val == "0":
            if fname.endswith(".pack"):
                mb.showerror("Wrong File Extension",
                             "⚠️  You've selected a wrong file!\n\n"
                             "FOTA Mode 0 (Full FOTA) requires a  .pac  file,\n"
                             "but the File Name ends with  .pack\n\n"
                             "Please update the File Name and try again.")
                return False
            if not fname.endswith(".pac"):
                mb.showerror("Wrong File Extension",
                             "⚠️  You've selected a wrong file!\n\n"
                             "FOTA Mode 0 (Full FOTA) requires a  .pac  file,\n"
                             "but the File Name does not end with  .pac\n\n"
                             "Please update the File Name and try again.")
                return False
        elif mode_val == "1":
            if fname.endswith(".pac") and not fname.endswith(".pack"):
                mb.showerror("Wrong File Extension",
                             "⚠️  You've selected a wrong file!\n\n"
                             "FOTA Mode 1 (Delta FOTA) requires a  .pack  file,\n"
                             "but the File Name ends with  .pac\n\n"
                             "Please update the File Name and try again.")
                return False
            if not fname.endswith(".pack"):
                mb.showerror("Wrong File Extension",
                             "⚠️  You've selected a wrong file!\n\n"
                             "FOTA Mode 1 (Delta FOTA) requires a  .pack  file,\n"
                             "but the File Name does not end with  .pack\n\n"
                             "Please update the File Name and try again.")
                return False
        return True

    def _run_with_retries(self, step_index, step_name, attempt_fn, max_attempts=3):
        popup = self._progress_popup
        result = None
        for attempt in range(1, max_attempts + 1):
            if popup._cancelled:
                return result if result is not None else (False, "Cancelled")
            if attempt > 1:
                self.after(0, self.log,
                           "🔁  '" + step_name + "' failed — retry " +
                           str(attempt) + "/" + str(max_attempts) + " …")
                self.after(0, popup.set_step_status, step_index, "running",
                           "Retry " + str(attempt) + "/" + str(max_attempts) + " …")
                time.sleep(2)
            try:
                result = attempt_fn()
            except Exception as e:
                result = (False, type(e).__name__ + ": " + str(e))
            if result[0]:
                return result
            self.after(0, self.log,
                       "⚠️  '" + step_name + "' attempt " + str(attempt) + "/" +
                       str(max_attempts) + " failed — " + str(result[1]))
        return result

    def _on_fota_nic_busy(self, step_name, reason):
        self._fota_running = False
        self._unfreeze_ui()
        self._set_status_pill("✘  NIC busy", "#a33025")
        self.status_var.set("✘  FOTA failed — NIC busy at: " + step_name)
        self.log("❌  FOTA stopped after 3 attempts at: " + step_name + " — " + reason)
        self.fota_btn.set_colors(DANGER, DANGER_HOVER)
        self.fota_btn.setText("❌  Failed — Retry")
        self.fota_btn.setEnabled(True)
        mb.showerror("NIC Busy",
                     "The FOTA update could not proceed because the NIC "
                     "did not respond after 3 attempts at:\n\n"
                     "🔸  " + step_name + "\n\n"
                     "Last error:\n" + str(reason) + "\n\n"
                     "The NIC is likely busy or unreachable right now. "
                     "Please wait a moment and try again.")

    # ── Send FOTA — orchestrates the full step sequence ─────────
    def send_fota(self):
        if self._fota_running:
            return
        if not self._validate_file_extension():
            return

        cmd = self.command_var.get()
        self._fota_running = True
        self._freeze_ui()
        self.fota_btn.set_colors(WARNING, WARNING_HOVER)
        self.fota_btn.setText("⏳  Updating …")
        self.fota_btn.setEnabled(False)
        self._set_status_pill("⏳  FOTA running", "#e6791f")
        self.status_var.set("⏳  FOTA sequence starting …")
        self.log("=" * 60)
        self.log("🚀  Starting full FOTA sequence …")
        self.log("    CMD → " + cmd)

        self._progress_popup = FotaStepDialog(self, self.FOTA_STEPS)
        self._progress_popup.show()
        threading.Thread(target=self._fota_worker, args=(cmd,), daemon=True).start()

    def _fota_worker(self, cmd: str):
        ip, port  = self._get_ip_port()
        file_path = self._get_file_path()
        payload   = (cmd + "\r\n").encode("ascii")
        popup     = self._progress_popup
        steps     = self.FOTA_STEPS

        def set_step(i, status, detail=None):
            self.after(0, popup.set_step_status, i, status, detail)

        def stop_failed(i, reason):
            self.after(0, popup.mark_failed, i, reason)
            self.after(0, self._on_fota_failed, steps[i], reason)

        def cancelled_mid_run():
            if popup._cancelled:
                self.after(0, popup.mark_cancelled)
                self.after(0, self._on_fota_cancelled)
                return True
            return False

        current_step = 0
        try:
            # ── Step 1: URL format ──
            current_step = 0
            set_step(0, "running")
            if cancelled_mid_run():
                return
            ok, msg = self._validate_url_format(file_path)
            if not ok:
                set_step(0, "fail", msg)
                stop_failed(0, msg)
                return
            set_step(0, "ok", msg)

            # ── Step 2: NIC IPv6 address valid ──
            current_step = 1
            set_step(1, "running")
            if cancelled_mid_run():
                return
            ok, msg = validate_ipv6_format(ip)
            if not ok:
                set_step(1, "fail", msg)
                stop_failed(1, msg)
                return
            set_step(1, "ok", msg)

            # ── Step 3: firmware file accessible (advisory) ──
            current_step = 2
            set_step(2, "running")
            if cancelled_mid_run():
                return
            ok, msg = self._check_remote_file_exists(file_path)
            if not ok:
                set_step(2, "warn", msg + "  (continuing — this check is informational only)")
            else:
                set_step(2, "ok", msg)

            # ── Step 4: ping test (advisory) ──
            current_step = 3
            set_step(3, "running")
            if cancelled_mid_run():
                return
            ok, msg = self._ping_test_ip(ip)
            if not ok:
                set_step(3, "warn", msg + "  (continuing — ICMP may be blocked even though TCP works)")
            else:
                set_step(3, "ok", msg)

            # ── Step 5: read current firmware version (retried x3) ──
            current_step = 4
            set_step(4, "running")
            if cancelled_mid_run():
                return
            ok, msg, old_ver = self._run_with_retries(
                4, "Read current firmware version",
                lambda: self._read_firmware_version_step(ip, port))
            if not ok:
                set_step(4, "fail", msg)
                self.after(0, popup.mark_failed, 4, msg)
                self.after(0, self._on_fota_nic_busy, steps[4], msg)
                return
            set_step(4, "ok", "Current version: " + old_ver)

            # ── Step 6: send FOTA command (retried x3) ──
            current_step = 5
            set_step(5, "running")
            if cancelled_mid_run():
                return
            ok, msg = self._run_with_retries(
                5, "Send FOTA command",
                lambda: self._send_fota_command_step(ip, port, payload))
            if not ok:
                set_step(5, "fail", msg)
                self.after(0, popup.mark_failed, 5, msg)
                self.after(0, self._on_fota_nic_busy, steps[5], msg)
                return
            set_step(5, "ok", msg)

            # ── Step 7: wait 30s ──
            current_step = 6
            set_step(6, "running")
            for remaining in range(30, 0, -1):
                if popup._cancelled:
                    self.after(0, popup.mark_cancelled)
                    self.after(0, self._on_fota_cancelled)
                    return
                self.after(0, popup.update_wait_countdown, 6, remaining)
                time.sleep(1)
            set_step(6, "ok", "30s wait complete")

            # ── Step 8: reread firmware version (retried x3, advisory) ──
            current_step = 7
            set_step(7, "running")
            ok, msg, new_ver = self._run_with_retries(
                7, "Reread firmware version",
                lambda: self._reread_firmware_version_step(ip, port))
            if not ok:
                set_step(7, "warn", str(msg) + "  (command was sent — verify version manually)")
                new_ver = "Unknown"
            else:
                set_step(7, "ok", "New version: " + new_ver)

            self.after(0, popup.mark_all_complete, old_ver, new_ver)
            self.after(0, self._save_current_settings)
            self.after(0, self._on_fota_send_ok, old_ver, new_ver)

        except Exception as e:
            reason = type(e).__name__ + ": " + str(e)
            self.after(0, popup.mark_failed, current_step, reason)
            self.after(0, self._on_fota_failed, steps[current_step], reason)

    # ── Step implementations ──────────────────────────────────
    def _read_firmware_version_step(self, ip, port):
        ok, msg, result = hdlc_read_firmware_blocking(
            ip, port, lambda m: self.after(0, self.log, m))
        if not ok:
            return False, msg, None
        fw = result.get("firmware")
        self.after(0, self.log, "📥  Firmware version: " + fw)
        self.after(0, self.version_var.set, fw)
        return True, "OK", fw

    def _send_fota_command_step(self, ip, port, payload):
        try:
            sa = resolve_ipv6_sockaddr(ip, port)
            self.after(0, self.log, "Connecting to " + sa[0] + " port " + str(sa[1]) + " …")
            with make_tcp_socket(timeout=10) as sock:
                sock.connect(sa)
                self.after(0, self.log, "Connected. Sending payload …")
                self.after(0, self.log, "RAW → " + repr(payload))
                sock.sendall(payload)
                self.after(0, self.log, "Sent " + str(len(payload)) + " bytes ✔")
                try:
                    resp = sock.recv(512).decode(errors="replace").strip()
                    if resp:
                        self.after(0, self.log, "NIC response: " + resp)
                except (socket.timeout, ConnectionResetError):
                    self.after(0, self.log,
                               "No immediate response — NIC is downloading firmware (normal)")
            return True, "FOTA command sent ✔"

        except ConnectionResetError:
            self.after(0, self.log,
                       "ℹ️  Connection closed by NIC (expected — NIC is rebooting to apply firmware)")
            return True, "FOTA command sent (NIC closed connection — expected)"

        except Exception as e:
            reason = type(e).__name__ + ": " + str(e)
            return False, reason

    def _reread_firmware_version_step(self, ip, port):
        ok, msg, result = hdlc_read_firmware_blocking(
            ip, port, lambda m: self.after(0, self.log, m))
        if not ok:
            return False, msg, None
        fw = result.get("firmware")
        self.after(0, self.log, "📥  Post-update firmware version: " + fw)
        self.after(0, self.version_var.set, fw)
        return True, "OK", fw

    # ── Completion / failure / cancellation handlers ────────────
    def _on_fota_send_ok(self, old_ver, new_ver):
        self._fota_running = False
        self._unfreeze_ui()
        self._set_status_pill("✅  Idle", "#1e7e34")
        self.status_var.set("✅  FOTA complete — " + (old_ver or "?") + " → " + (new_ver or "?"))
        self.log("✅  FOTA sequence fully complete.")
        self.fota_btn.set_colors(BRAND, BRAND_HOVER)
        self.fota_btn.setText("🚀  Send FOTA")
        self.fota_btn.setEnabled(True)
        mb.showinfo("FOTA Update Complete",
                     "Firmware update finished successfully!\n\n"
                     "Previous version: " + (old_ver or "unknown") + "\n"
                     "New version: " + (new_ver or "unknown"))

    def _on_fota_failed(self, step_name, reason):
        self._fota_running = False
        self._unfreeze_ui()
        self._set_status_pill("✘  FOTA failed", "#a33025")
        self.status_var.set("✘  FOTA failed — " + step_name)
        self.log("❌  FOTA stopped at: " + step_name + " — " + reason)
        self.fota_btn.set_colors(DANGER, DANGER_HOVER)
        self.fota_btn.setText("❌  Failed — Retry")
        self.fota_btn.setEnabled(True)
        mb.showerror("FOTA Update Blocked",
                     "The FOTA update process was stopped at:\n\n"
                     "🔸  " + step_name + "\n\n"
                     "Reason:\n" + reason + "\n\n"
                     "Please resolve the issue and try again.")

    def _on_fota_cancelled(self):
        self._fota_running = False
        self._unfreeze_ui()
        self._set_status_pill("⚠  Cancelled", "#b8860b")
        self.status_var.set("⚠  FOTA cancelled by user")
        self.log("⚠  FOTA cancelled by user.")
        self.fota_btn.set_colors(WARNING, WARNING_HOVER)
        self.fota_btn.setText("🚀  Send FOTA")
        self.fota_btn.setEnabled(True)


def main():
    app = QtWidgets.QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setFont(qfont(FONT_FAMILY, 10))
    win = OrionFOTAApp()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()